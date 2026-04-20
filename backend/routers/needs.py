"""
Needs Router — Liste de besoins (gérante / admin).

Périmètre :
  - Liste de TOUS les besoins (salle, salle_jeux, jardin, cuisine, toilettes, autres).
  - Gérante crée un besoin (prix optionnel, multi-items), Admin approuve et
    peut convertir automatiquement en demande d'achats (dépense pending).
  - Endpoint /needs/analysis réutilise la logique de forecasts pour :
    doublons (14j, demandes + achats stock), intra-doublons, stock matches,
    redundant items, recent purchases et impact trésorerie.

Statuts : en_attente | traite | annule
"""
from fastapi import APIRouter, HTTPException, Body, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field, ConfigDict
from datetime import datetime, timezone, timedelta
from typing import List, Optional
import io
import uuid
import logging

from .forecasts import analyze_single_request, _compute_treasury

try:
    from services.sms_service import send_admin_sms_notification
except Exception:  # pragma: no cover
    async def send_admin_sms_notification(_msg: str) -> bool:
        return False

router = APIRouter(tags=["needs"])
db = None
logger = logging.getLogger(__name__)


def set_db(database):
    global db
    db = database


# ==================== MODELS ====================

VALID_LOCATIONS = {"salle", "salle_jeux", "jardin", "cuisine", "toilettes", "autres"}
VALID_STATUSES = {"en_attente", "traite", "annule"}
VALID_URGENCY = {"normale", "urgente"}


class NeedItem(BaseModel):
    model_config = ConfigDict(extra="ignore")
    location: str = "autres"  # salle | salle_jeux | jardin | cuisine | toilettes | autres
    description: str
    quantity: int = 1
    unit_price: Optional[float] = 0
    amount: Optional[float] = 0
    notes: Optional[str] = ""


class NeedCreate(BaseModel):
    location: str = "autres"
    description: str
    items: Optional[List[NeedItem]] = None
    quantity: Optional[int] = 1
    unit_price: Optional[float] = 0
    amount: Optional[float] = 0
    supplier: Optional[str] = None
    urgency: str = "normale"
    notes: Optional[str] = ""
    requested_by: str


class NeedUpdate(BaseModel):
    location: Optional[str] = None
    description: Optional[str] = None
    items: Optional[List[NeedItem]] = None
    quantity: Optional[int] = None
    unit_price: Optional[float] = None
    amount: Optional[float] = None
    supplier: Optional[str] = None
    urgency: Optional[str] = None
    notes: Optional[str] = None
    status: Optional[str] = None
    admin_notes: Optional[str] = None


# ==================== CRUD ====================

@router.get("/needs")
async def list_needs(status: Optional[str] = None, location: Optional[str] = None):
    """List needs. Optional filters: status, location."""
    try:
        query = {}
        if status:
            query["status"] = status
        if location:
            query["location"] = location
        items = await db.needs.find(query, {"_id": 0}).sort("created_at", -1).to_list(2000)
        return {"needs": items}
    except Exception as e:
        logger.error(f"Error listing needs: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/needs")
async def create_need(data: NeedCreate):
    try:
        if data.location not in VALID_LOCATIONS:
            data.location = "autres"
        if data.urgency not in VALID_URGENCY:
            data.urgency = "normale"

        doc = {
            "id": str(uuid.uuid4()),
            "location": data.location,
            "description": data.description,
            "items": [it.model_dump() for it in (data.items or [])],
            "quantity": data.quantity or 1,
            "unit_price": data.unit_price or 0,
            "amount": data.amount or 0,
            "supplier": data.supplier,
            "urgency": data.urgency,
            "notes": data.notes or "",
            "requested_by": data.requested_by,
            "status": "en_attente",
            "created_at": datetime.now(timezone.utc).isoformat(),
        }
        await db.needs.insert_one(doc)

        # Admin notification on EVERY new need (urgent highlighted specifically)
        try:
            loc_labels = {
                "salle": "Salle", "salle_jeux": "Salle de jeux",
                "jardin": "Jardin", "cuisine": "Cuisine",
                "toilettes": "Toilettes", "autres": "Autres",
            }
            is_urgent = doc["urgency"] == "urgente"
            items_lines = []
            for it in doc.get("items") or []:
                q = it.get("quantity") or 1
                desc = (it.get("description") or "").strip()[:40]
                if desc:
                    items_lines.append(f"- {desc} x{q}")
            items_block = "\n".join(items_lines[:6]) or "(sans article detaille)"
            extra_count = max(0, len(doc.get("items") or []) - 6)
            extra = f"\n+ {extra_count} autre(s)..." if extra_count > 0 else ""

            header = "[URGENT] " if is_urgent else "[BESOIN] "
            msg = (
                f"{header}Nouveau besoin Espace Maxo\n"
                f"Espace: {loc_labels.get(doc.get('location',''), doc.get('location',''))}\n"
                f"Demande: {doc.get('description','')[:80]}\n"
                f"Par: {doc.get('requested_by','-')}\n"
                f"Articles ({len(doc.get('items') or [])}):\n"
                f"{items_block}{extra}"
            )
            if doc.get("amount"):
                msg += f"\nMontant estime: {doc['amount']:,.0f} F".replace(",", " ")
            await send_admin_sms_notification(msg)
        except Exception as notif_err:
            logger.error(f"Admin notification failed: {notif_err}")

        return {"success": True, "need": {k: v for k, v in doc.items() if k != "_id"}}
    except Exception as e:
        logger.error(f"Error creating need: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/needs/{need_id}")
async def update_need(need_id: str, data: NeedUpdate):
    try:
        update = {k: v for k, v in data.model_dump(exclude_unset=True).items() if v is not None}
        if "location" in update and update["location"] not in VALID_LOCATIONS:
            update["location"] = "autres"
        if "status" in update and update["status"] not in VALID_STATUSES:
            update["status"] = "en_attente"
        if "urgency" in update and update["urgency"] not in VALID_URGENCY:
            update["urgency"] = "normale"
        if "items" in update and update["items"] is not None:
            update["items"] = [
                it if isinstance(it, dict) else it.model_dump()
                for it in update["items"]
            ]
        update["updated_at"] = datetime.now(timezone.utc).isoformat()
        res = await db.needs.update_one({"id": need_id}, {"$set": update})
        if res.matched_count == 0:
            raise HTTPException(status_code=404, detail="Besoin non trouvé")
        return {"success": True}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating need: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/needs/{need_id}")
async def delete_need(need_id: str):
    try:
        res = await db.needs.delete_one({"id": need_id})
        if res.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Besoin non trouvé")
        return {"success": True}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting need: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== ADMIN ACTIONS ====================

@router.post("/needs/{need_id}/cancel")
async def cancel_need(need_id: str, body: dict = Body(default={})):
    """Admin cancels (annule) a need."""
    try:
        reason = (body or {}).get("reason") or ""
        res = await db.needs.update_one(
            {"id": need_id},
            {"$set": {
                "status": "annule",
                "admin_notes": reason,
                "cancelled_at": datetime.now(timezone.utc).isoformat(),
            }}
        )
        if res.matched_count == 0:
            raise HTTPException(status_code=404, detail="Besoin non trouvé")
        return {"success": True}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error cancelling need: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/needs/{need_id}/convert-to-expense")
async def convert_need_to_expense(need_id: str, body: dict = Body(default={})):
    """Admin: convertit un besoin en demande d'achats (expense pending) et marque le besoin 'traite'.

    Le body peut contenir: category (défaut "autres"), overrides unit_price/amount si gérante n'a pas renseigné.
    """
    try:
        need = await db.needs.find_one({"id": need_id}, {"_id": 0})
        if not need:
            raise HTTPException(status_code=404, detail="Besoin non trouvé")
        if need.get("status") == "traite":
            raise HTTPException(status_code=400, detail="Besoin déjà traité")

        category = (body or {}).get("category") or "autres"
        # Build expense items list from need.items or fallback single item
        src_items = need.get("items") or []
        exp_items = []
        total_amount = 0
        if src_items:
            for it in src_items:
                qty = it.get("quantity", 1) or 1
                up = it.get("unit_price", 0) or 0
                amt = it.get("amount") or (qty * up)
                exp_items.append({
                    "category": category,
                    "description": it.get("description") or "",
                    "quantity": qty,
                    "unit_price": up,
                    "amount": amt,
                })
                total_amount += amt
        else:
            qty = need.get("quantity", 1) or 1
            up = need.get("unit_price", 0) or 0
            amt = need.get("amount") or (qty * up)
            exp_items.append({
                "category": category,
                "description": need.get("description") or "",
                "quantity": qty,
                "unit_price": up,
                "amount": amt,
            })
            total_amount = amt

        expense_doc = {
            "id": str(uuid.uuid4()),
            "category": category,
            "description": f"Besoin {need.get('location', 'autres')} - {need.get('description') or ''}",
            "quantity": len(exp_items),
            "unit_price": None,
            "amount": total_amount,
            "supplier": need.get("supplier"),
            "planned_date": None,
            "receipt_image": None,
            "requested_by": need.get("requested_by"),
            "is_group": True,
            "group_id": need.get("id"),
            "items": exp_items,
            "status": "pending",
            "source_need_id": need_id,
            "created_at": datetime.now(timezone.utc).isoformat(),
        }
        await db.expenses.insert_one(expense_doc)

        await db.needs.update_one(
            {"id": need_id},
            {"$set": {
                "status": "traite",
                "converted_to_expense_id": expense_doc["id"],
                "converted_at": datetime.now(timezone.utc).isoformat(),
            }}
        )
        return {
            "success": True,
            "expense_id": expense_doc["id"],
            "expense": {k: v for k, v in expense_doc.items() if k != "_id"},
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error converting need: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== ANALYSIS ====================

@router.get("/needs/analysis")
async def needs_analysis():
    """Analyse chaque besoin 'en_attente' contre :
    - Les autres besoins récents (14j)
    - Les demandes d'achats récentes (expenses 14j)
    - Les achats stock réels (stock_purchases 14j)
    - Le stock actuel
    - La trésorerie disponible
    """
    try:
        today = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
        lookback = today - timedelta(days=14)
        lookback_str = lookback.strftime("%Y-%m-%d")

        treasury = await _compute_treasury(today)
        available = treasury["available"]

        needs_pending = await db.needs.find({"status": "en_attente"}, {"_id": 0}).to_list(500)
        recent_needs = await db.needs.find({"created_at": {"$gte": lookback_str}}, {"_id": 0}).to_list(1000)
        recent_expenses = await db.expenses.find({"created_at": {"$gte": lookback_str}}, {"_id": 0}).to_list(1000)
        # Merge both as "recent_requests"
        recent_requests = list(recent_needs) + list(recent_expenses)

        stock_products = await db.stock_products.find({"is_active": True}, {"_id": 0}).to_list(2000)
        recent_purchases = await db.stock_purchases.find(
            {"created_at": {"$gte": lookback_str}}, {"_id": 0}
        ).sort("created_at", -1).to_list(500)

        analyses = []
        for n in needs_pending:
            analyses.append(await analyze_single_request(
                db, n, recent_requests, recent_purchases, stock_products, available,
                id_field="id", self_ref="need_id",
            ))
        return {"treasury": treasury, "analyses": analyses}
    except Exception as e:
        logger.error(f"Error building needs analysis: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ==================== EXPORTS (PDF / EXCEL) ====================

LOCATION_LABELS = {
    "salle": "Salle",
    "salle_jeux": "Salle de jeux",
    "jardin": "Jardin",
    "cuisine": "Cuisine",
    "toilettes": "Toilettes",
    "autres": "Autres",
}
STATUS_LABELS = {
    "en_attente": "En attente",
    "traite": "Traité",
    "annule": "Annulé",
}


async def _fetch_filtered_needs(status: Optional[str], location: Optional[str],
                                date_from: Optional[str], date_to: Optional[str]):
    query = {}
    if status and status != "all":
        query["status"] = status
    if location and location != "all":
        query["location"] = location
    if date_from or date_to:
        rng = {}
        if date_from:
            rng["$gte"] = date_from
        if date_to:
            rng["$lte"] = date_to + "T23:59:59"
        query["created_at"] = rng
    return await db.needs.find(query, {"_id": 0}).sort("created_at", -1).to_list(5000)


@router.get("/needs/export/pdf")
async def export_needs_pdf(
    status: Optional[str] = None,
    location: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
):
    """Export the filtered needs list as a PDF document."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

        needs = await _fetch_filtered_needs(status, location, date_from, date_to)

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer, pagesize=A4,
            topMargin=15*mm, bottomMargin=15*mm, leftMargin=12*mm, rightMargin=12*mm,
        )
        styles = getSampleStyleSheet()
        elements = []

        title_style = ParagraphStyle(
            'NeedsTitle', parent=styles['Heading1'], fontSize=16,
            textColor=colors.HexColor('#4c51bf'), alignment=1,
        )
        elements.append(Paragraph("Liste de besoins — Espace Maxo", title_style))
        elements.append(Spacer(1, 4*mm))

        filter_parts = []
        if status and status != "all":
            filter_parts.append(f"Statut : {STATUS_LABELS.get(status, status)}")
        if location and location != "all":
            filter_parts.append(f"Espace : {LOCATION_LABELS.get(location, location)}")
        if date_from:
            filter_parts.append(f"Du : {date_from}")
        if date_to:
            filter_parts.append(f"Au : {date_to}")
        filter_parts.append(f"Total : {len(needs)} besoin(s)")
        filter_style = ParagraphStyle(
            'NeedsFilters', parent=styles['Normal'], fontSize=9,
            textColor=colors.grey, alignment=1,
        )
        elements.append(Paragraph(" | ".join(filter_parts), filter_style))
        elements.append(Spacer(1, 6*mm))

        # Summary KPI
        total_items = sum(len(n.get("items") or []) for n in needs)
        total_amount = sum(n.get("amount", 0) or 0 for n in needs)
        by_status = {}
        for n in needs:
            k = n.get("status", "en_attente")
            by_status[k] = by_status.get(k, 0) + 1
        kpi_row = [
            ["Besoins", "Articles", "Montant total"],
            [str(len(needs)), str(total_items), f"{total_amount:,.0f} F".replace(",", " ")],
        ]
        kpi = Table(kpi_row, colWidths=[60*mm, 60*mm, 60*mm])
        kpi.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4c51bf')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f7fafc')),
        ]))
        elements.append(kpi)
        elements.append(Spacer(1, 6*mm))

        # Main table : one row per need, then collapsed items line
        header = ["Date", "Espace", "Description", "Articles", "Montant", "Urgence", "Statut", "Par"]
        data = [header]
        for n in needs[:300]:
            created = (n.get("created_at") or "")[:10]
            data.append([
                created,
                LOCATION_LABELS.get(n.get("location", "autres"), n.get("location", "")),
                (n.get("description", "") or "")[:40],
                str(len(n.get("items") or [])),
                f"{(n.get('amount', 0) or 0):,.0f}".replace(",", " "),
                "Urgent" if n.get("urgency") == "urgente" else "Normale",
                STATUS_LABELS.get(n.get("status", "en_attente"), n.get("status", "")),
                (n.get("requested_by", "") or "")[:18],
            ])

        col_w = [18*mm, 22*mm, 48*mm, 14*mm, 22*mm, 16*mm, 18*mm, 28*mm]
        table = Table(data, colWidths=col_w, repeatRows=1)
        style_commands = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a202c')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('ALIGN', (3, 0), (4, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#e2e8f0')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f7fafc')]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ]
        # Urgency coloring
        for i, n in enumerate(needs[:300], 1):
            if n.get("urgency") == "urgente":
                style_commands.append(('TEXTCOLOR', (5, i), (5, i), colors.HexColor('#e53e3e')))
            st = n.get("status", "en_attente")
            st_color = {
                "en_attente": colors.HexColor('#d69e2e'),
                "traite": colors.HexColor('#2f855a'),
                "annule": colors.HexColor('#c53030'),
            }.get(st)
            if st_color:
                style_commands.append(('TEXTCOLOR', (6, i), (6, i), st_color))
        table.setStyle(TableStyle(style_commands))
        elements.append(table)
        elements.append(Spacer(1, 6*mm))

        # Detailed items per need
        detail_title = ParagraphStyle(
            'DetailTitle', parent=styles['Heading3'], fontSize=11,
            textColor=colors.HexColor('#4c51bf'),
        )
        elements.append(Paragraph("Détail des articles", detail_title))
        elements.append(Spacer(1, 3*mm))
        for n in needs[:100]:
            items = n.get("items") or []
            if not items:
                continue
            sub_title_style = ParagraphStyle(
                'SubTitle', parent=styles['Normal'], fontSize=9,
                textColor=colors.HexColor('#2d3748'),
            )
            elements.append(Paragraph(
                f"<b>{LOCATION_LABELS.get(n.get('location','autres'),'')}</b> — "
                f"{n.get('description','')} <font color='grey'>({(n.get('created_at') or '')[:10]})</font>",
                sub_title_style,
            ))
            sub_data = [["#", "Article", "Espace", "Qté", "P.U.", "Total"]]
            for idx, it in enumerate(items, 1):
                qty = it.get("quantity", 1) or 1
                up = it.get("unit_price", 0) or 0
                amt = it.get("amount") or qty * up
                sub_data.append([
                    str(idx),
                    (it.get("description", "") or "")[:36],
                    LOCATION_LABELS.get(it.get("location", "autres"), ""),
                    str(qty),
                    f"{up:,.0f}".replace(",", " "),
                    f"{amt:,.0f}".replace(",", " "),
                ])
            sub_w = [10*mm, 70*mm, 28*mm, 14*mm, 22*mm, 28*mm]
            sub_t = Table(sub_data, colWidths=sub_w, repeatRows=1)
            sub_t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e2e8f0')),
                ('FONTSIZE', (0, 0), (-1, -1), 7),
                ('GRID', (0, 0), (-1, -1), 0.2, colors.HexColor('#cbd5e0')),
                ('ALIGN', (3, 0), (5, -1), 'RIGHT'),
                ('TOPPADDING', (0, 0), (-1, -1), 1),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
            ]))
            elements.append(sub_t)
            elements.append(Spacer(1, 3*mm))

        doc.build(elements)
        buffer.seek(0)
        filename = f"liste_besoins_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M')}.pdf"
        return StreamingResponse(
            buffer, media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    except Exception as e:
        logger.error(f"Error exporting needs PDF: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/needs/export/excel")
async def export_needs_excel(
    status: Optional[str] = None,
    location: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
):
    """Export the filtered needs list as an Excel workbook (2 sheets)."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        needs = await _fetch_filtered_needs(status, location, date_from, date_to)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Besoins"

        thin = Side(style="thin", color="CBD5E0")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_fill = PatternFill("solid", fgColor="4C51BF")
        header_align = Alignment(horizontal="center", vertical="center")

        # Title row
        ws.merge_cells("A1:H1")
        ws["A1"] = "Liste de besoins — Espace Maxo"
        ws["A1"].font = Font(bold=True, size=14, color="4C51BF")
        ws["A1"].alignment = Alignment(horizontal="center")

        # Filters row
        ws.merge_cells("A2:H2")
        filter_parts = []
        if status and status != "all":
            filter_parts.append(f"Statut: {STATUS_LABELS.get(status, status)}")
        if location and location != "all":
            filter_parts.append(f"Espace: {LOCATION_LABELS.get(location, location)}")
        if date_from:
            filter_parts.append(f"Du: {date_from}")
        if date_to:
            filter_parts.append(f"Au: {date_to}")
        filter_parts.append(f"Total: {len(needs)}")
        ws["A2"] = " | ".join(filter_parts)
        ws["A2"].font = Font(italic=True, color="718096", size=9)
        ws["A2"].alignment = Alignment(horizontal="center")

        headers = ["Date", "Espace", "Description", "Articles", "Montant", "Urgence", "Statut", "Par"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=4, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = header_align
            c.border = border

        row = 5
        for n in needs:
            ws.cell(row=row, column=1, value=(n.get("created_at") or "")[:10])
            ws.cell(row=row, column=2, value=LOCATION_LABELS.get(n.get("location", "autres"), ""))
            ws.cell(row=row, column=3, value=n.get("description", "") or "")
            ws.cell(row=row, column=4, value=len(n.get("items") or []))
            ws.cell(row=row, column=5, value=n.get("amount", 0) or 0)
            urg_cell = ws.cell(row=row, column=6, value="Urgent" if n.get("urgency") == "urgente" else "Normale")
            if n.get("urgency") == "urgente":
                urg_cell.font = Font(bold=True, color="E53E3E")
            st = n.get("status", "en_attente")
            st_cell = ws.cell(row=row, column=7, value=STATUS_LABELS.get(st, st))
            st_cell.font = Font(
                color={"en_attente": "D69E2E", "traite": "2F855A", "annule": "C53030"}.get(st, "000000"),
                bold=True,
            )
            ws.cell(row=row, column=8, value=n.get("requested_by", "") or "")
            for col in range(1, 9):
                ws.cell(row=row, column=col).border = border
            row += 1

        # Auto-width
        for col_idx in range(1, 9):
            letter = openpyxl.utils.get_column_letter(col_idx)
            max_len = 10
            for r in range(4, row):
                v = ws.cell(row=r, column=col_idx).value
                if v and not isinstance(ws.cell(row=r, column=col_idx), openpyxl.cell.cell.MergedCell):
                    max_len = max(max_len, min(50, len(str(v)) + 2))
            ws.column_dimensions[letter].width = max_len

        # Sheet 2 : items per need
        ws2 = wb.create_sheet("Articles détaillés")
        ws2.merge_cells("A1:G1")
        ws2["A1"] = "Détail des articles"
        ws2["A1"].font = Font(bold=True, size=12, color="4C51BF")
        ws2["A1"].alignment = Alignment(horizontal="center")
        headers2 = ["Date", "Besoin", "Espace article", "Article", "Quantité", "Prix unitaire", "Total"]
        for col, h in enumerate(headers2, 1):
            c = ws2.cell(row=3, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = header_align
            c.border = border

        r2 = 4
        for n in needs:
            nd_desc = n.get("description", "") or ""
            n_date = (n.get("created_at") or "")[:10]
            for it in (n.get("items") or []):
                qty = it.get("quantity", 1) or 1
                up = it.get("unit_price", 0) or 0
                amt = it.get("amount") or qty * up
                ws2.cell(row=r2, column=1, value=n_date)
                ws2.cell(row=r2, column=2, value=nd_desc)
                ws2.cell(row=r2, column=3, value=LOCATION_LABELS.get(it.get("location", "autres"), ""))
                ws2.cell(row=r2, column=4, value=it.get("description", "") or "")
                ws2.cell(row=r2, column=5, value=qty)
                ws2.cell(row=r2, column=6, value=up)
                ws2.cell(row=r2, column=7, value=amt)
                for col in range(1, 8):
                    ws2.cell(row=r2, column=col).border = border
                r2 += 1
        for col_idx in range(1, 8):
            letter = openpyxl.utils.get_column_letter(col_idx)
            max_len = 10
            for r in range(3, r2):
                v = ws2.cell(row=r, column=col_idx).value
                if v and not isinstance(ws2.cell(row=r, column=col_idx), openpyxl.cell.cell.MergedCell):
                    max_len = max(max_len, min(50, len(str(v)) + 2))
            ws2.column_dimensions[letter].width = max_len

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        filename = f"liste_besoins_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M')}.xlsx"
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    except Exception as e:
        logger.error(f"Error exporting needs Excel: {e}")
        raise HTTPException(status_code=500, detail=str(e))
