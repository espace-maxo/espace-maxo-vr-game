from fastapi import APIRouter, HTTPException, Body, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from typing import List, Optional, Dict
from datetime import datetime, timezone, timedelta
import logging
import uuid
import random
import re
import bcrypt
import io

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/stock", tags=["stock"])
db = None

def set_db(database):
    global db
    db = database

# ==================== AUTH MODELS ====================

class StockUserCreate(BaseModel):
    username: str
    password: str
    full_name: str = ""
    role: str = "consultation"  # administrateur, gerant, magasinier, consultation

class StockLoginRequest(BaseModel):
    username: str
    password: str

# ==================== AUTH ENDPOINTS ====================

@router.post("/auth/login")
async def stock_login(data: StockLoginRequest):
    user = await db.stock_users.find_one({"username": data.username, "is_active": True})
    if not user:
        raise HTTPException(401, "Identifiants incorrects")
    
    if not bcrypt.checkpw(data.password.encode('utf-8'), user["password_hash"].encode('utf-8')):
        raise HTTPException(401, "Identifiants incorrects")
    
    # Update last login
    await db.stock_users.update_one({"id": user["id"]}, {"$set": {"last_login": datetime.now(timezone.utc).isoformat()}})
    
    return {
        "success": True,
        "user": {
            "id": user["id"],
            "username": user["username"],
            "full_name": user["full_name"],
            "role": user["role"]
        }
    }

@router.get("/auth/users")
async def get_stock_users():
    users = await db.stock_users.find({}, {"_id": 0, "password_hash": 0}).sort("full_name", 1).to_list(100)
    return {"users": users}

@router.post("/auth/users")
async def create_stock_user(data: StockUserCreate):
    existing = await db.stock_users.find_one({"username": data.username})
    if existing:
        raise HTTPException(400, f"Le nom d'utilisateur '{data.username}' existe deja")
    
    password_hash = bcrypt.hashpw(data.password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
    
    user = {
        "id": str(uuid.uuid4()),
        "username": data.username,
        "password_hash": password_hash,
        "full_name": data.full_name or data.username,
        "role": data.role,
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "last_login": None
    }
    await db.stock_users.insert_one(user)
    user.pop("_id", None)
    user.pop("password_hash", None)
    return {"success": True, "user": user}

@router.put("/auth/users/{user_id}")
async def update_stock_user(user_id: str, data: dict = Body(...)):
    data.pop("_id", None)
    data.pop("id", None)
    data.pop("password_hash", None)
    # If password is being changed
    if "password" in data and data["password"]:
        data["password_hash"] = bcrypt.hashpw(data.pop("password").encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
    else:
        data.pop("password", None)
    await db.stock_users.update_one({"id": user_id}, {"$set": data})
    updated = await db.stock_users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})
    return {"success": True, "user": updated}

@router.delete("/auth/users/{user_id}")
async def delete_stock_user(user_id: str):
    await db.stock_users.delete_one({"id": user_id})
    return {"success": True}

@router.post("/auth/seed-users")
async def seed_stock_users():
    """Create default stock users if none exist"""
    existing = await db.stock_users.count_documents({})
    if existing > 0:
        return {"success": True, "message": f"{existing} utilisateurs deja presents"}
    
    default_users = [
        {"username": "admin", "password": "Admin2026", "full_name": "Administrateur", "role": "administrateur"},
        {"username": "gerante", "password": "Gerante2026", "full_name": "Gerante", "role": "gerant"},
        {"username": "magasinier", "password": "Magasin2026", "full_name": "Magasinier", "role": "magasinier"},
        {"username": "consultation", "password": "Consult2026", "full_name": "Consultation", "role": "consultation"},
    ]
    
    for u in default_users:
        pw_hash = bcrypt.hashpw(u["password"].encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        await db.stock_users.insert_one({
            "id": str(uuid.uuid4()),
            "username": u["username"],
            "password_hash": pw_hash,
            "full_name": u["full_name"],
            "role": u["role"],
            "is_active": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "last_login": None
        })
    
    return {"success": True, "message": f"{len(default_users)} utilisateurs crees"}

# ==================== MODELS ====================

class StockCategoryCreate(BaseModel):
    name: str
    description: str = ""
    color: str = "#3b82f6"
    icon: str = "Package"
    subcategories: list = []

class StockProductCreate(BaseModel):
    code: str = ""
    name: str
    category_id: str
    subcategory: str = ""
    unit: str = "kg"
    quantity: float = 0
    stock_min: float = 5
    stock_max: float = 100
    purchase_price: float = 0
    sale_price: float = 0  # Prix de vente unitaire (pour valoriser le stock en valeur de revente)
    supplier_id: str = ""
    storage_location: str = ""
    # 'cuisine' (default): auto-déstockage via factures/recettes
    # 'magasin'         : stock isolé, déstockage MANUEL uniquement (aucune synchro Caisse)
    storage_zone: str = "cuisine"
    is_active: bool = True
    photo_url: str = ""
    date_achat: str = ""
    date_peremption: str = ""
    observation: str = ""

class StockMovementCreate(BaseModel):
    product_id: str
    movement_type: str  # entree, sortie, ajustement, perte, casse, retour_fournisseur, inventaire
    quantity: float
    unit_price: float = 0
    reason: str = ""
    user_name: str = ""

class StockPurchaseCreate(BaseModel):
    supplier_id: str = ""
    supplier_name: str = ""
    purchase_date: str = ""
    items: List[Dict] = []  # [{product_id, product_name, quantity, unit_price}]
    notes: str = ""
    user_name: str = ""

class StockSupplierCreate(BaseModel):
    name: str
    phone: str = ""
    email: str = ""
    address: str = ""
    product_types: str = ""
    notes: str = ""


# Portionnement rules: governs how purchase quantities are converted into "portions" stock units.
# Liquid categories (boissons, huiles) are non-portionnable: 1 unit = 1 unit (no conversion).
# Non-liquid categories: portion_factor (e.g., 5 portions/kg) defines conversion.
class PortionCategoryRule(BaseModel):
    category_id: str
    portions_per_unit: float = 1.0  # 1.0 = no conversion (1 unit purchased = 1 portion stocked)
    is_liquid: bool = False  # If True: stays in original unit, no conversion ever.

class PortionProductOverride(BaseModel):
    stock_product_id: str
    portions_per_unit: float = 1.0  # Override the category rule for this specific product
    is_liquid: Optional[bool] = None  # Override liquid flag (None = inherit from category)
    purchase_unit: Optional[str] = None  # Purchase unit (kg, l, piece, pot, etc.) — saved on the stock product
    daily_consumption: float = 0.0  # Time-based deduction: portions consumed per day (0 = none)

class PortionRulesUpdate(BaseModel):
    category_rules: List[PortionCategoryRule] = []
    product_overrides: List[PortionProductOverride] = []


# ==================== CATEGORIES ====================

@router.get("/categories")
async def get_categories():
    cats = await db.stock_categories.find({}, {"_id": 0}).sort("name", 1).to_list(100)
    return {"categories": cats}

@router.post("/categories")
async def create_category(data: StockCategoryCreate):
    cat = {
        "id": str(uuid.uuid4()),
        "name": data.name,
        "description": data.description,
        "color": data.color,
        "icon": data.icon,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.stock_categories.insert_one(cat)
    cat.pop("_id", None)
    return {"success": True, "category": cat}

@router.put("/categories/{cat_id}")
async def update_category(cat_id: str, data: dict = Body(...)):
    await db.stock_categories.update_one({"id": cat_id}, {"$set": data})
    updated = await db.stock_categories.find_one({"id": cat_id}, {"_id": 0})
    return {"success": True, "category": updated}

@router.delete("/categories/{cat_id}")
async def delete_category(cat_id: str):
    count = await db.stock_products.count_documents({"category_id": cat_id})
    if count > 0:
        raise HTTPException(400, f"Impossible de supprimer : {count} produit(s) dans cette catégorie")
    await db.stock_categories.delete_one({"id": cat_id})
    return {"success": True}

# ==================== PRODUCTS ====================

@router.get("/products")
async def get_products(category_id: str = None, status: str = None, search: str = None, alert: str = None, storage_zone: str = None):
    query = {}
    if category_id:
        query["category_id"] = category_id
    if status == "active":
        query["is_active"] = True
    elif status == "inactive":
        query["is_active"] = False
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"code": {"$regex": search, "$options": "i"}}
        ]
    if storage_zone == "magasin":
        query["storage_zone"] = "magasin"
    elif storage_zone == "cuisine":
        # Default cuisine = everything that is NOT explicitly magasin (back-compat for old rows)
        query["storage_zone"] = {"$ne": "magasin"}
    
    products = await db.stock_products.find(query, {"_id": 0}).sort("name", 1).to_list(2000)
    
    if alert == "rupture":
        products = [p for p in products if p.get("quantity", 0) <= 0]
    elif alert == "faible":
        products = [p for p in products if 0 < p.get("quantity", 0) <= p.get("stock_min", 5)]
    
    return {"products": products}

@router.get("/products/{product_id}")
async def get_product(product_id: str):
    p = await db.stock_products.find_one({"id": product_id}, {"_id": 0})
    if not p:
        raise HTTPException(404, "Produit non trouvé")
    return p

@router.post("/products")
async def create_product(data: StockProductCreate):
    if data.code:
        existing = await db.stock_products.find_one({"code": data.code})
        if existing:
            raise HTTPException(400, f"Le code produit '{data.code}' existe déjà")
    
    product = {
        "id": str(uuid.uuid4()),
        "code": data.code or f"PRD-{str(uuid.uuid4())[:6].upper()}",
        "name": data.name,
        "category_id": data.category_id,
        "subcategory": data.subcategory,
        "unit": data.unit,
        "quantity": data.quantity,
        "stock_min": data.stock_min,
        "stock_max": data.stock_max,
        "purchase_price": data.purchase_price,
        "sale_price": data.sale_price,
        "valeur_stock": data.quantity * data.purchase_price,
        "valeur_stock_vente": data.quantity * data.sale_price,
        "supplier_id": data.supplier_id,
        "storage_location": data.storage_location,
        "storage_zone": data.storage_zone or "cuisine",
        "is_active": data.is_active,
        "photo_url": data.photo_url,
        "date_achat": data.date_achat,
        "date_peremption": data.date_peremption,
        "observation": data.observation,
        "statut": "rupture" if data.quantity <= 0 else ("faible" if data.quantity <= data.stock_min else "normal"),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    await db.stock_products.insert_one(product)
    product.pop("_id", None)
    return {"success": True, "product": product}

@router.put("/products/{product_id}")
async def update_product(product_id: str, data: dict = Body(...)):
    data.pop("_id", None)
    data.pop("id", None)
    data["updated_at"] = datetime.now(timezone.utc).isoformat()
    # Recalculate valeur_stock and statut
    product = await db.stock_products.find_one({"id": product_id})
    if not product:
        raise HTTPException(404, "Produit non trouve")
    qty = data.get("quantity", product.get("quantity", 0))
    price = data.get("purchase_price", product.get("purchase_price", 0))
    sale_price = data.get("sale_price", product.get("sale_price", 0))
    smin = data.get("stock_min", product.get("stock_min", 5))
    data["valeur_stock"] = qty * price
    data["valeur_stock_vente"] = qty * sale_price
    data["statut"] = "rupture" if qty <= 0 else ("faible" if qty <= smin else "normal")
    await db.stock_products.update_one({"id": product_id}, {"$set": data})
    updated = await db.stock_products.find_one({"id": product_id}, {"_id": 0})
    return {"success": True, "product": updated}

@router.delete("/products/{product_id}")
async def delete_product(product_id: str):
    await db.stock_products.delete_one({"id": product_id})
    return {"success": True}

# ==================== MOVEMENTS ====================

@router.get("/movements")
async def get_movements(product_id: str = None, movement_type: str = None, date_from: str = None, date_to: str = None, limit: int = 100):
    query = {}
    if product_id:
        query["product_id"] = product_id
    if movement_type:
        query["movement_type"] = movement_type
    if date_from or date_to:
        date_q = {}
        if date_from:
            date_q["$gte"] = date_from
        if date_to:
            date_q["$lte"] = date_to + "T23:59:59"
        query["created_at"] = date_q
    
    movements = await db.stock_movements.find(query, {"_id": 0}).sort("created_at", -1).to_list(limit)
    return {"movements": movements}


# ==================== STOCK À UNE DATE DONNÉE (BOISSONS) ====================

def _is_drinks_category(cat: dict) -> bool:
    """Détecte si une catégorie de stock concerne les boissons / bar."""
    name = (cat or {}).get("name") or ""
    n = name.lower()
    return ("boisson" in n) or ("bar" in n) or ("cocktail" in n)


@router.get("/snapshot")
async def stock_snapshot_at(
    at: str = "",
    only_drinks: bool = True,
):
    """Reconstruit le stock de chaque produit à une date+heure donnée.

    Logique : `qty_at(t) = current_quantity - somme(delta des mouvements après t)`.
    Pour chaque mouvement, le delta net est `new_quantity - previous_quantity`.

    Paramètres :
      - at : date ISO. Si vide ou si format `YYYY-MM-DD`, utilise 23:59:59 ce jour-là.
      - only_drinks : si True (défaut), filtre sur catégories de type Boisson/Bar/Cocktail.
    """
    if not at:
        raise HTTPException(422, "Paramètre 'at' requis (YYYY-MM-DD ou ISO datetime)")
    # Normalise en bornant à fin de journée si seule la date est fournie.
    cutoff = at if "T" in at else f"{at}T23:59:59.999"

    # Catégories boissons ?
    drink_cat_ids = set()
    if only_drinks:
        cats = await db.stock_categories.find({}, {"_id": 0}).to_list(500)
        drink_cat_ids = {c["id"] for c in cats if _is_drinks_category(c)}

    # Produits à inclure
    prod_query = {}
    if only_drinks:
        if not drink_cat_ids:
            return {"at": cutoff, "only_drinks": True, "products": [], "total_products": 0,
                    "total_quantity": 0.0, "total_value": 0.0}
        prod_query["category_id"] = {"$in": list(drink_cat_ids)}
    products = await db.stock_products.find(prod_query, {"_id": 0}).to_list(5000)

    # Cats lookup pour libellés
    cats_all = await db.stock_categories.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(500)
    cat_map = {c["id"]: c["name"] for c in cats_all}

    if not products:
        return {"at": cutoff, "only_drinks": only_drinks, "products": [], "total_products": 0,
                "total_quantity": 0.0, "total_value": 0.0}

    pids = [p["id"] for p in products]
    # Mouvements après la date (à soustraire du stock actuel)
    later_movs = await db.stock_movements.find(
        {"product_id": {"$in": pids}, "created_at": {"$gt": cutoff}},
        {"_id": 0, "product_id": 1, "previous_quantity": 1, "new_quantity": 1, "quantity": 1, "movement_type": 1},
    ).to_list(100000)

    # Net delta after cutoff per product
    delta_after: Dict[str, float] = {}
    for m in later_movs:
        pid = m.get("product_id")
        if not pid:
            continue
        prev_q = m.get("previous_quantity")
        new_q = m.get("new_quantity")
        if prev_q is not None and new_q is not None:
            delta = float(new_q) - float(prev_q)
        else:
            # Fallback (mouvements anciens sans previous/new)
            qty = float(m.get("quantity") or 0)
            mtype = (m.get("movement_type") or "").lower()
            if mtype in ("entree", "entrée", "in"): delta = qty
            elif mtype in ("sortie", "out"): delta = -qty
            else: delta = 0.0
        delta_after[pid] = delta_after.get(pid, 0.0) + delta

    rows = []
    total_qty = 0.0
    total_val = 0.0
    for p in products:
        pid = p["id"]
        cur_qty = float(p.get("quantity") or 0)
        qty_at = cur_qty - delta_after.get(pid, 0.0)
        if qty_at < 0:
            qty_at = 0.0  # garde-fou pour anomalies historiques
        unit_cost = float(p.get("purchase_price") or 0)
        value_at = qty_at * unit_cost
        rows.append({
            "id": pid,
            "code": p.get("code") or "",
            "name": p.get("name") or "—",
            "category_id": p.get("category_id"),
            "category_name": cat_map.get(p.get("category_id"), "—"),
            "subcategory": p.get("subcategory") or "",
            "unit": p.get("unit") or "",
            "current_quantity": cur_qty,
            "quantity_at": round(qty_at, 4),
            "delta_after": round(delta_after.get(pid, 0.0), 4),
            "unit_purchase_price": unit_cost,
            "value_at": round(value_at, 2),
            "stock_min": float(p.get("stock_min") or 0),
            "movements_after_count": sum(1 for m in later_movs if m.get("product_id") == pid),
        })
        total_qty += qty_at
        total_val += value_at

    rows.sort(key=lambda r: (r["category_name"] or "", r["name"] or ""))
    return {
        "at": cutoff,
        "only_drinks": only_drinks,
        "drink_categories": [{"id": cid, "name": cat_map.get(cid, "—")} for cid in drink_cat_ids],
        "total_products": len(rows),
        "total_quantity": round(total_qty, 2),
        "total_value": round(total_val, 2),
        "products": rows,
    }


@router.get("/destock-live")
async def destock_live_dashboard(limit: int = 50):
    """Live deduction dashboard.

    Returns:
      - recent_sales: last `limit` stock movements that originated from a sale
        (movement.invoice_id present). Joined with caisse product name when available.
      - linked_count, total_caisse_count: ratio of linked caisse products.
      - unlinked_caisse_products: caisse products with NO stock_product_id (they
        never destock, the user can link them in Produits Caisse).
      - linked_no_sales: caisse products linked but with no sale in the last 30
        days (suspected misconfig OR slow-mover).
    """
    now = datetime.now(timezone.utc)
    cutoff_iso = (now - timedelta(days=30)).isoformat()

    # Auto-trigger daily deductions (silent, idempotent: only applies if elapsed days >= 1)
    try:
        await _apply_daily_deductions_internal(silent=True)
    except Exception as e:
        logger.warning(f"Daily deduction auto-trigger failed: {e}")

    # Recent sales (movements tied to invoices)
    recent_sales = await db.stock_movements.find(
        {"invoice_id": {"$exists": True, "$ne": None}, "movement_type": "sortie"},
        {"_id": 0},
    ).sort("created_at", -1).to_list(limit)

    # Caisse products linkage state — uses NEW stock_links array (multi-link), with legacy fallback to stock_product_id.
    # Products marked as "no_stock_tracking" (services/games/fees) are excluded from the linked/unlinked stats entirely.
    caisse_products = await db.caisse_products.find({}, {"_id": 0}).to_list(2000)
    trackable_products = [cp for cp in caisse_products if not cp.get("no_stock_tracking")]
    service_products = [cp for cp in caisse_products if cp.get("no_stock_tracking")]
    total_caisse = len(trackable_products)
    def _is_linked(cp):
        return bool(cp.get("stock_links")) or bool(cp.get("stock_product_id")) or bool(cp.get("stock_recipe_id"))
    linked = [cp for cp in trackable_products if _is_linked(cp)]
    unlinked = [cp for cp in trackable_products if not _is_linked(cp)]

    # For linked products, find last sale movement (using caisse_product_id)
    sales_by_caisse_id = {}
    movs_with_caisse = await db.stock_movements.find(
        {"caisse_product_id": {"$exists": True, "$ne": None}, "movement_type": "sortie"},
        {"_id": 0, "caisse_product_id": 1, "created_at": 1, "quantity": 1},
    ).sort("created_at", -1).to_list(5000)
    for m in movs_with_caisse:
        cp_id = m.get("caisse_product_id")
        if cp_id and cp_id not in sales_by_caisse_id:
            sales_by_caisse_id[cp_id] = m.get("created_at")

    linked_no_sales = []
    linked_with_sales = []
    for cp in linked:
        last_sale = sales_by_caisse_id.get(cp.get("id"))
        # Resolve effective stock links (multi → list, legacy single → wrap)
        effective_links = cp.get("stock_links") or ([cp["stock_product_id"]] if cp.get("stock_product_id") else [])
        info = {
            "id": cp.get("id"),
            "name": cp.get("name"),
            "category": cp.get("category"),
            "stock_product_id": cp.get("stock_product_id"),  # legacy field, kept for backwards compat
            "stock_links": effective_links,
            "stock_recipe_id": cp.get("stock_recipe_id") or "",
            "last_sale_at": last_sale,
        }
        if not last_sale or last_sale < cutoff_iso:
            linked_no_sales.append(info)
        else:
            linked_with_sales.append(info)

    return {
        "recent_sales": recent_sales,
        "summary": {
            "total_caisse_products": total_caisse,
            "linked_count": len(linked),
            "unlinked_count": len(unlinked),
            "linked_no_sales_count": len(linked_no_sales),
            "linked_with_recent_sales_count": len(linked_with_sales),
            "recent_sales_count": len(recent_sales),
            "service_products_count": len(service_products),
        },
        "unlinked_caisse_products": [
            {"id": cp.get("id"), "name": cp.get("name"), "category": cp.get("category"), "price": cp.get("price")}
            for cp in unlinked
        ],
        "linked_no_sales": linked_no_sales,
    }


@router.get("/products/{product_id}/analysis")
async def product_period_analysis(product_id: str, start_date: str, end_date: str):
    """Analyse entrées/sorties d'un produit sur une période donnée.

    Retourne :
    - Solde avant période (calculé à rebours sur l'historique)
    - Total entrées / sorties / pertes / ajustements / inventaires
    - Solde théorique (début + entrées - sorties - pertes ± ajustements)
    - Solde réel actuel (depuis stock_products)
    - Écart (théorique vs réel)
    - Détection d'anomalies : gaspillage, sur-consommation, écart inexpliqué
    - Breakdown quotidien
    """
    try:
        product = await db.stock_products.find_one({"id": product_id}, {"_id": 0})
        if not product:
            raise HTTPException(status_code=404, detail="Produit introuvable")

        # Validate dates
        try:
            datetime.strptime(start_date, "%Y-%m-%d")
            datetime.strptime(end_date, "%Y-%m-%d")
        except ValueError:
            raise HTTPException(status_code=400, detail="Dates invalides (format YYYY-MM-DD requis)")

        start_iso = start_date + "T00:00:00"
        end_iso = end_date + "T23:59:59"

        # All movements during the period
        movements_period = await db.stock_movements.find(
            {"product_id": product_id, "created_at": {"$gte": start_iso, "$lte": end_iso}},
            {"_id": 0}
        ).sort("created_at", 1).to_list(5000)

        # Movements BEFORE period (to compute opening balance)
        # opening_balance = current_quantity - (sum of all movement deltas after beginning of period)
        # Simplification: use the first movement in the period's `previous_quantity` if available,
        # else fallback to current product quantity minus period net change.
        opening_balance = None
        if movements_period:
            first = movements_period[0]
            if "previous_quantity" in first:
                opening_balance = float(first.get("previous_quantity", 0) or 0)
        current_quantity = float(product.get("quantity", 0) or 0)

        # Aggregation by type
        def _qty_of(m):
            return float(m.get("quantity", 0) or 0)

        totals = {
            "entree": 0.0, "sortie": 0.0, "perte": 0.0, "casse": 0.0,
            "retour_fournisseur": 0.0, "ajustement_positif": 0.0,
            "ajustement_negatif": 0.0, "inventaire": 0.0,
            "transfert_entree": 0.0, "transfert_sortie": 0.0,
        }
        sorties_detail = {"auto_facture": 0.0, "manuel": 0.0, "transfert": 0.0, "autre_sortie": 0.0}
        for m in movements_period:
            mt = m.get("movement_type", "")
            q = _qty_of(m)
            if mt == "ajustement":
                # Distinguer positif/négatif via delta previous → new
                new_q = float(m.get("new_quantity", 0) or 0)
                prev_q = float(m.get("previous_quantity", 0) or 0)
                if new_q >= prev_q:
                    totals["ajustement_positif"] += (new_q - prev_q)
                else:
                    totals["ajustement_negatif"] += (prev_q - new_q)
            elif mt in totals:
                totals[mt] += q
            # Breakdown des sorties
            if mt == "sortie":
                reason = (m.get("reason") or "").lower()
                if "facture" in reason or "auto" in reason or m.get("invoice_id"):
                    sorties_detail["auto_facture"] += q
                elif "transfert" in reason:
                    sorties_detail["transfert"] += q
                elif m.get("user_name"):
                    sorties_detail["manuel"] += q
                else:
                    sorties_detail["autre_sortie"] += q

        total_entrees = totals["entree"] + totals["retour_fournisseur"] + totals["transfert_entree"] + totals["ajustement_positif"]
        total_sorties = totals["sortie"] + totals["perte"] + totals["casse"] + totals["transfert_sortie"] + totals["ajustement_negatif"]

        # Solde théorique (approximation : dernier mouvement reflète la réalité via new_quantity)
        closing_theorical = None
        closing_real_end = None
        if movements_period:
            closing_real_end = float(movements_period[-1].get("new_quantity", 0) or 0)
            if opening_balance is not None:
                closing_theorical = opening_balance + total_entrees - total_sorties
        else:
            # Aucun mouvement sur la période
            opening_balance = current_quantity
            closing_theorical = current_quantity
            closing_real_end = current_quantity

        ecart = None
        if closing_theorical is not None and closing_real_end is not None:
            ecart = round(closing_real_end - closing_theorical, 3)

        # Détection anomalies
        anomalies = []
        severity = "ok"  # ok | warning | critical

        # 1. Écart inexpliqué entre théorique et réel
        if ecart is not None and abs(ecart) > 0.1:
            anomalies.append({
                "type": "ecart_stock",
                "severity": "warning" if abs(ecart) < max(1, total_entrees * 0.05) else "critical",
                "message": f"Écart de {ecart:+.2f} {product.get('unit', '')} entre solde théorique et réel",
                "value": ecart,
            })

        # 2. Pertes / casse élevées
        pertes_total = totals["perte"] + totals["casse"]
        if pertes_total > 0 and total_entrees > 0:
            ratio_pertes = pertes_total / total_entrees
            if ratio_pertes > 0.15:
                anomalies.append({
                    "type": "pertes_elevees",
                    "severity": "critical",
                    "message": f"Pertes/casses représentent {ratio_pertes*100:.1f}% des entrées ({pertes_total:.2f} perdu(s))",
                    "value": ratio_pertes,
                })
            elif ratio_pertes > 0.05:
                anomalies.append({
                    "type": "pertes_moderees",
                    "severity": "warning",
                    "message": f"Pertes/casses à {ratio_pertes*100:.1f}% des entrées",
                    "value": ratio_pertes,
                })

        # 3. Sortie sans entrée correspondante (stock négatif évité)
        if total_sorties > (opening_balance or 0) + total_entrees + 0.01:
            anomalies.append({
                "type": "sorties_sans_couverture",
                "severity": "critical",
                "message": "Sorties supérieures au stock disponible (entrées + solde initial) — incohérence",
                "value": total_sorties - ((opening_balance or 0) + total_entrees),
            })

        # 4. Rupture actuelle
        stock_min = float(product.get("stock_min", 5) or 0)
        if current_quantity <= 0:
            anomalies.append({"type": "rupture", "severity": "critical", "message": "Produit en rupture de stock actuellement", "value": current_quantity})
        elif current_quantity <= stock_min:
            anomalies.append({"type": "stock_faible", "severity": "warning", "message": f"Stock actuel ({current_quantity}) sous le minimum ({stock_min})", "value": current_quantity})

        # 5. Aucun mouvement sur la période = produit "dormant"
        if len(movements_period) == 0:
            anomalies.append({
                "type": "produit_dormant",
                "severity": "warning",
                "message": "Aucun mouvement sur la période — produit inactif",
                "value": 0,
            })

        # Severity globale
        if any(a["severity"] == "critical" for a in anomalies):
            severity = "critical"
        elif any(a["severity"] == "warning" for a in anomalies):
            severity = "warning"

        # Breakdown quotidien (simple)
        from collections import defaultdict
        daily = defaultdict(lambda: {"date": "", "entrees": 0.0, "sorties": 0.0, "net": 0.0})
        for m in movements_period:
            d = (m.get("created_at") or "")[:10]
            daily[d]["date"] = d
            q = _qty_of(m)
            mt = m.get("movement_type", "")
            if mt in ("entree", "retour_fournisseur", "transfert_entree"):
                daily[d]["entrees"] += q
                daily[d]["net"] += q
            elif mt in ("sortie", "perte", "casse", "transfert_sortie"):
                daily[d]["sorties"] += q
                daily[d]["net"] -= q
            elif mt == "ajustement":
                delta = float(m.get("new_quantity", 0) or 0) - float(m.get("previous_quantity", 0) or 0)
                if delta >= 0:
                    daily[d]["entrees"] += delta
                else:
                    daily[d]["sorties"] += abs(delta)
                daily[d]["net"] += delta
        daily_list = sorted(daily.values(), key=lambda x: x["date"])

        return {
            "product": {
                "id": product.get("id"),
                "name": product.get("name"),
                "code": product.get("code", ""),
                "unit": product.get("unit", ""),
                "stock_min": stock_min,
                "current_quantity": current_quantity,
                "storage_zone": product.get("storage_zone", "cuisine"),
            },
            "period": {"start_date": start_date, "end_date": end_date, "movements_count": len(movements_period)},
            "balance": {
                "opening": round(opening_balance or 0, 3),
                "current": round(current_quantity, 3),
                "theorical_at_end": round(closing_theorical or 0, 3) if closing_theorical is not None else None,
                "real_at_end": round(closing_real_end or 0, 3) if closing_real_end is not None else None,
                "ecart": ecart,
            },
            "totals": {k: round(v, 3) for k, v in totals.items()},
            "total_entrees": round(total_entrees, 3),
            "total_sorties": round(total_sorties, 3),
            "net_movement": round(total_entrees - total_sorties, 3),
            "sorties_breakdown": {k: round(v, 3) for k, v in sorties_detail.items()},
            "anomalies": anomalies,
            "severity": severity,
            "daily": daily_list,
            "movements": movements_period[-50:],  # Derniers 50 mouvements
        }
    except HTTPException:
        raise
    except Exception as e:
        import traceback; traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


    cutoff_iso = (now - timedelta(days=30)).isoformat()

    # Auto-trigger daily deductions (silent, idempotent: only applies if elapsed days >= 1)
    try:
        await _apply_daily_deductions_internal(silent=True)
    except Exception as e:
        logger.warning(f"Daily deduction auto-trigger failed: {e}")

    # Recent sales (movements tied to invoices)
    recent_sales = await db.stock_movements.find(
        {"invoice_id": {"$exists": True, "$ne": None}, "movement_type": "sortie"},
        {"_id": 0},
    ).sort("created_at", -1).to_list(limit)

    # Caisse products linkage state — uses NEW stock_links array (multi-link), with legacy fallback to stock_product_id.
    # Products marked as "no_stock_tracking" (services/games/fees) are excluded from the linked/unlinked stats entirely.
    caisse_products = await db.caisse_products.find({}, {"_id": 0}).to_list(2000)
    trackable_products = [cp for cp in caisse_products if not cp.get("no_stock_tracking")]
    service_products = [cp for cp in caisse_products if cp.get("no_stock_tracking")]
    total_caisse = len(trackable_products)
    def _is_linked(cp):
        return bool(cp.get("stock_links")) or bool(cp.get("stock_product_id")) or bool(cp.get("stock_recipe_id"))
    linked = [cp for cp in trackable_products if _is_linked(cp)]
    unlinked = [cp for cp in trackable_products if not _is_linked(cp)]

    # For linked products, find last sale movement (using caisse_product_id)
    sales_by_caisse_id = {}
    movs_with_caisse = await db.stock_movements.find(
        {"caisse_product_id": {"$exists": True, "$ne": None}, "movement_type": "sortie"},
        {"_id": 0, "caisse_product_id": 1, "created_at": 1, "quantity": 1},
    ).sort("created_at", -1).to_list(5000)
    for m in movs_with_caisse:
        cp_id = m.get("caisse_product_id")
        if cp_id and cp_id not in sales_by_caisse_id:
            sales_by_caisse_id[cp_id] = m.get("created_at")

    linked_no_sales = []
    linked_with_sales = []
    for cp in linked:
        last_sale = sales_by_caisse_id.get(cp.get("id"))
        # Resolve effective stock links (multi → list, legacy single → wrap)
        effective_links = cp.get("stock_links") or ([cp["stock_product_id"]] if cp.get("stock_product_id") else [])
        info = {
            "id": cp.get("id"),
            "name": cp.get("name"),
            "category": cp.get("category"),
            "stock_product_id": cp.get("stock_product_id"),  # legacy field, kept for backwards compat
            "stock_links": effective_links,
            "stock_recipe_id": cp.get("stock_recipe_id") or "",
            "last_sale_at": last_sale,
        }
        if not last_sale or last_sale < cutoff_iso:
            linked_no_sales.append(info)
        else:
            linked_with_sales.append(info)

    return {
        "recent_sales": recent_sales,
        "summary": {
            "total_caisse_products": total_caisse,
            "linked_count": len(linked),
            "unlinked_count": len(unlinked),
            "linked_no_sales_count": len(linked_no_sales),
            "linked_with_recent_sales_count": len(linked_with_sales),
            "recent_sales_count": len(recent_sales),
            "service_products_count": len(service_products),
        },
        "unlinked_caisse_products": [
            {"id": cp.get("id"), "name": cp.get("name"), "category": cp.get("category"), "price": cp.get("price")}
            for cp in unlinked
        ],
        "linked_no_sales": linked_no_sales,
    }


@router.get("/links-overview")
async def caisse_stock_links_overview():
    """Bi-directional view of Caisse↔Stock links.

    Returns:
      - caisse_to_stock: list of caisse products with their resolved stock_links (multi)
        and resolved stock product names.
      - stock_to_caisse: list of stock products with all caisse products that link to them
        (computed by inverse traversal).
      - recipes: list of caisse products linked via a recipe (separate, not part of multi-link).
    """
    caisse_products = await db.caisse_products.find({}, {"_id": 0}).to_list(5000)
    stock_products = await db.stock_products.find({"is_active": True}, {"_id": 0}).to_list(5000)
    sp_by_id = {sp.get("id"): sp for sp in stock_products}

    caisse_to_stock = []
    recipes_view = []
    services_view = []  # Caisse products marked as no_stock_tracking
    stock_to_caisse_map = {}  # stock_id -> list of caisse {id, name, category}

    for cp in caisse_products:
        cp_id = cp.get("id")
        cp_name = cp.get("name", "")
        cp_cat = cp.get("category", "")
        cp_dept = cp.get("department", "")
        # Service products (no_stock_tracking=True) are listed separately, never in unlinked.
        if cp.get("no_stock_tracking"):
            services_view.append({
                "caisse_id": cp_id,
                "caisse_name": cp_name,
                "category": cp_cat,
                "department": cp_dept,
            })
            continue
        # Resolve effective links: prefer stock_links (multi), fallback legacy single
        link_ids = list(cp.get("stock_links") or [])
        if not link_ids and cp.get("stock_product_id"):
            link_ids = [cp["stock_product_id"]]
        recipe_id = cp.get("stock_recipe_id") or ""

        if recipe_id:
            recipes_view.append({
                "caisse_id": cp_id,
                "caisse_name": cp_name,
                "category": cp_cat,
                "department": cp_dept,
                "recipe_id": recipe_id,
            })
            continue

        resolved_links = []
        for sid in link_ids:
            sp = sp_by_id.get(sid)
            if sp:
                resolved_links.append({
                    "stock_id": sid,
                    "stock_name": sp.get("name", ""),
                    "stock_code": sp.get("code", ""),
                    "current_quantity": sp.get("quantity", 0),
                    "unit": sp.get("unit", ""),
                })
                stock_to_caisse_map.setdefault(sid, []).append({
                    "caisse_id": cp_id,
                    "caisse_name": cp_name,
                    "category": cp_cat,
                    "department": cp_dept,
                })

        caisse_to_stock.append({
            "caisse_id": cp_id,
            "caisse_name": cp_name,
            "category": cp_cat,
            "department": cp_dept,
            "links": resolved_links,
            "links_count": len(resolved_links),
        })

    stock_to_caisse = []
    for sp in stock_products:
        sid = sp.get("id")
        consumers = stock_to_caisse_map.get(sid, [])
        stock_to_caisse.append({
            "stock_id": sid,
            "stock_name": sp.get("name", ""),
            "stock_code": sp.get("code", ""),
            "current_quantity": sp.get("quantity", 0),
            "unit": sp.get("unit", ""),
            "consumers": consumers,
            "consumers_count": len(consumers),
        })

    # Sort: most-linked first for stock view; alpha for caisse view
    stock_to_caisse.sort(key=lambda x: (-x["consumers_count"], x["stock_name"]))
    caisse_to_stock.sort(key=lambda x: (-x["links_count"], x["caisse_name"]))

    return {
        "caisse_to_stock": caisse_to_stock,
        "stock_to_caisse": stock_to_caisse,
        "recipes": recipes_view,
        "services": services_view,
        "summary": {
            "total_caisse_products": len(caisse_products),
            "trackable_caisse_products": len(caisse_products) - len(services_view),
            "total_stock_products": len(stock_products),
            "caisse_with_multi_links": sum(1 for x in caisse_to_stock if x["links_count"] > 1),
            "caisse_with_links": sum(1 for x in caisse_to_stock if x["links_count"] >= 1),
            "caisse_with_recipe": len(recipes_view),
            "caisse_services": len(services_view),
            "stock_with_consumers": sum(1 for x in stock_to_caisse if x["consumers_count"] >= 1),
        },
    }



# ==================== PORTIONNEMENT RULES ====================

# Default liquid category names (case-insensitive substring match)
LIQUID_CATEGORY_KEYWORDS = [
    "boisson", "boissons", "huile", "huiles", "matiere grasse", "matieres grasses",
    "cocktail", "bar", "sirop",
]

def _is_liquid_category(category_name: str) -> bool:
    n = (category_name or "").lower()
    return any(kw in n for kw in LIQUID_CATEGORY_KEYWORDS)


@router.get("/portionnement/rules")
async def get_portionnement_rules():
    """Get all portionnement rules: product-level only.
    Returns the full list of active stock products with their effective rule.
    For products without an override:
      - portions_per_unit defaults to 1.0
      - is_liquid is auto-detected from the category name (boisson/huile/cocktail).
    """
    products = await db.stock_products.find({"is_active": True}, {"_id": 0}).sort("name", 1).to_list(5000)
    cats = await db.stock_categories.find({}, {"_id": 0}).to_list(500)
    cat_by_id = {c.get("id"): c for c in cats}

    saved_prod_rules = {r["stock_product_id"]: r async for r in db.portion_product_overrides.find({}, {"_id": 0})}

    rules_resp = []
    for sp in products:
        spid = sp.get("id")
        cat = cat_by_id.get(sp.get("category_id", ""))
        cat_name = (cat or {}).get("name", "")
        saved = saved_prod_rules.get(spid)
        # Default liquid detection from category name (UI suggestion only)
        default_liquid = _is_liquid_category(cat_name)
        if saved:
            ppu = float(saved.get("portions_per_unit", 1.0) or 1.0)
            is_liq = saved.get("is_liquid")
            if is_liq is None:
                is_liq = default_liquid
            else:
                is_liq = bool(is_liq)
            saved_pu = saved.get("purchase_unit") or sp.get("purchase_unit") or sp.get("unit", "")
            daily = float(saved.get("daily_consumption", 0) or 0)
        else:
            ppu = 1.0
            is_liq = default_liquid
            saved_pu = sp.get("purchase_unit") or sp.get("unit", "")
            daily = 0.0
        rules_resp.append({
            "stock_product_id": spid,
            "stock_product_name": sp.get("name", ""),
            "stock_product_code": sp.get("code", ""),
            "category_id": sp.get("category_id", ""),
            "category_name": cat_name,
            "current_unit": sp.get("unit", ""),
            "purchase_unit": saved_pu,
            "current_quantity": sp.get("quantity", 0),
            "portions_per_unit": ppu,
            "is_liquid": is_liq,
            "daily_consumption": daily,
            "last_daily_deduction_date": sp.get("last_daily_deduction_date", ""),
            "configured": saved is not None,
        })

    return {"product_rules": rules_resp}


@router.put("/portionnement/rules")
async def update_portionnement_rules(data: PortionRulesUpdate):
    """Replace product-level portionnement rules atomically.
    Also updates `purchase_unit` on each stock product so future BC use the right unit.
    Note: category-level rules are no longer supported (kept in payload for backwards compat but ignored).
    """
    now = datetime.now(timezone.utc).isoformat()
    await db.portion_product_overrides.delete_many({})
    if data.product_overrides:
        await db.portion_product_overrides.insert_many([
            {**r.model_dump(), "updated_at": now} for r in data.product_overrides
        ])
        # Persist purchase_unit on each stock product (only when provided)
        for r in data.product_overrides:
            if r.purchase_unit:
                await db.stock_products.update_one(
                    {"id": r.stock_product_id},
                    {"$set": {"purchase_unit": r.purchase_unit, "updated_at": now}}
                )
    # Drop legacy category rules: per-product is the new source of truth
    await db.portion_category_rules.delete_many({})
    return {"success": True, "rules_count": len(data.product_overrides)}


async def _resolve_portion_factor(stock_product: dict) -> tuple:
    """Returns (portions_per_unit: float, is_liquid: bool) for a given stock product.
    Priority: product rule → auto-detection by category name (default 1.0, is_liquid=False).
    """
    spid = stock_product.get("id")
    category_id = stock_product.get("category_id", "")

    override = await db.portion_product_overrides.find_one({"stock_product_id": spid}, {"_id": 0})
    if override is not None:
        ppu = float(override.get("portions_per_unit", 1.0) or 1.0)
        is_liq = override.get("is_liquid")
        if is_liq is None:
            cat = await db.stock_categories.find_one({"id": category_id}, {"_id": 0}) if category_id else None
            is_liq = _is_liquid_category((cat or {}).get("name", ""))
        return ppu, bool(is_liq)

    # No rule: default factor 1.0, auto-detect liquid by category name
    cat = await db.stock_categories.find_one({"id": category_id}, {"_id": 0}) if category_id else None
    is_liq = _is_liquid_category((cat or {}).get("name", ""))
    return 1.0, is_liq


@router.post("/portionnement/apply-units")
async def apply_portionnement_units():
    """One-shot migration: switch every non-liquid product unit to 'portion'.
    Liquid products keep their original unit. Quantities are NOT modified (per user choice)."""
    products = await db.stock_products.find({"is_active": True}, {"_id": 0}).to_list(5000)
    updated = 0
    skipped_liquid = 0
    for p in products:
        ppu, is_liq = await _resolve_portion_factor(p)
        if is_liq:
            skipped_liquid += 1
            continue
        # Only update if not already 'portion'
        if p.get("unit") != "portion":
            await db.stock_products.update_one(
                {"id": p["id"]},
                {"$set": {"unit": "portion", "purchase_unit": p.get("unit", ""), "updated_at": datetime.now(timezone.utc).isoformat()}}
            )
            updated += 1
    return {"success": True, "updated_to_portion": updated, "kept_liquid": skipped_liquid}



async def _apply_daily_deductions_internal(silent: bool = True) -> dict:
    """Run daily deduction for every product with daily_consumption > 0.
    Computes elapsed full days since last_daily_deduction_date (or override.updated_at),
    deducts (days * daily_consumption), stamps last_daily_deduction_date to today.
    Logs each deduction in stock_movements.
    Returns a summary report.
    """
    today = datetime.now(timezone.utc).date()
    today_iso = today.isoformat()
    now_iso = datetime.now(timezone.utc).isoformat()

    overrides = await db.portion_product_overrides.find(
        {"daily_consumption": {"$gt": 0}}, {"_id": 0}
    ).to_list(2000)
    if not overrides:
        return {"applied_count": 0, "total_deducted": 0, "details": []}

    details = []
    applied_count = 0
    total_deducted = 0.0

    for ov in overrides:
        spid = ov.get("stock_product_id")
        daily = float(ov.get("daily_consumption", 0) or 0)
        if daily <= 0:
            continue
        sp = await db.stock_products.find_one({"id": spid}, {"_id": 0})
        if not sp or not sp.get("is_active", True):
            continue
        # Skip "magasin" products — manual-only zone
        if sp.get("storage_zone") == "magasin":
            continue

        # Determine reference date: last deduction or override creation date (fall back to today)
        last_date_str = sp.get("last_daily_deduction_date") or ov.get("updated_at") or today_iso
        try:
            last_date = datetime.fromisoformat(last_date_str.replace("Z", "+00:00")).date()
        except Exception:
            last_date = today
        days_elapsed = (today - last_date).days
        if days_elapsed <= 0:
            continue

        deducted_qty = daily * days_elapsed
        old_qty = sp.get("quantity", 0)
        new_qty = max(0, old_qty - deducted_qty)
        actually_deducted = old_qty - new_qty
        smin = sp.get("stock_min", 5)
        new_statut = "rupture" if new_qty <= 0 else ("faible" if new_qty <= smin else "normal")
        new_valeur = new_qty * sp.get("purchase_price", 0)

        await db.stock_movements.insert_one({
            "id": str(uuid.uuid4()),
            "product_id": spid,
            "product_name": sp.get("name", ""),
            "product_code": sp.get("code", ""),
            "movement_type": "sortie",
            "quantity": actually_deducted,
            "previous_quantity": old_qty,
            "new_quantity": new_qty,
            "unit": sp.get("unit", ""),
            "unit_price": sp.get("purchase_price", 0),
            "total_value": actually_deducted * sp.get("purchase_price", 0),
            "reason": (
                f"Conso journalière auto ({daily}/jour × {days_elapsed} jour{'s' if days_elapsed > 1 else ''})"
            ),
            "user_name": "Système (auto)",
            "created_at": now_iso,
        })
        await db.stock_products.update_one(
            {"id": spid},
            {"$set": {
                "quantity": new_qty,
                "valeur_stock": new_valeur,
                "statut": new_statut,
                "last_daily_deduction_date": today_iso,
                "updated_at": now_iso,
            }}
        )
        applied_count += 1
        total_deducted += actually_deducted
        details.append({
            "stock_product_id": spid,
            "name": sp.get("name", ""),
            "days": days_elapsed,
            "daily": daily,
            "deducted": actually_deducted,
            "previous_qty": old_qty,
            "new_qty": new_qty,
        })

    if not silent:
        logger.info(f"Daily deduction: {applied_count} products, total {total_deducted} portions deducted")

    return {
        "applied_count": applied_count,
        "total_deducted": total_deducted,
        "details": details,
    }


@router.post("/portionnement/apply-daily")
async def apply_daily_deductions():
    """Manually trigger the daily deduction for all configured products."""
    return await _apply_daily_deductions_internal(silent=False)



@router.post("/movements")
async def create_movement(data: StockMovementCreate):
    product = await db.stock_products.find_one({"id": data.product_id})
    if not product:
        raise HTTPException(404, "Produit non trouvé")
    
    current_qty = product.get("quantity", 0)
    
    # Calculate new quantity
    if data.movement_type in ["entree", "retour_fournisseur", "transfert_entree"]:
        new_qty = current_qty + data.quantity
    elif data.movement_type in ["sortie", "perte", "casse", "transfert_sortie"]:
        if data.quantity > current_qty:
            raise HTTPException(400, f"Stock insuffisant. Disponible: {current_qty} {product.get('unit', '')}")
        new_qty = current_qty - data.quantity
    elif data.movement_type == "ajustement":
        new_qty = data.quantity  # Direct set
    elif data.movement_type == "inventaire":
        new_qty = data.quantity  # Set to physical count
    else:
        raise HTTPException(400, f"Type de mouvement invalide: {data.movement_type}")
    
    movement = {
        "id": str(uuid.uuid4()),
        "product_id": data.product_id,
        "product_name": product.get("name", ""),
        "product_code": product.get("code", ""),
        "movement_type": data.movement_type,
        "quantity": data.quantity,
        "previous_quantity": current_qty,
        "new_quantity": new_qty,
        "unit": product.get("unit", ""),
        "unit_price": data.unit_price,
        "total_value": data.quantity * data.unit_price,
        "reason": data.reason,
        "user_name": data.user_name,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.stock_movements.insert_one(movement)
    movement.pop("_id", None)
    
    # Update product quantity + valeur_stock + statut
    new_valeur = new_qty * product.get("purchase_price", 0)
    new_statut = "rupture" if new_qty <= 0 else ("faible" if new_qty <= product.get("stock_min", 5) else "normal")
    await db.stock_products.update_one(
        {"id": data.product_id},
        {"$set": {"quantity": new_qty, "valeur_stock": new_valeur, "statut": new_statut, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    return {"success": True, "movement": movement}

@router.delete("/movements/{movement_id}")
async def delete_movement(movement_id: str):
    result = await db.stock_movements.delete_one({"id": movement_id})
    if result.deleted_count == 0:
        raise HTTPException(404, "Mouvement non trouve")
    return {"success": True}

@router.post("/movements/delete-bulk")
async def delete_movements_bulk(ids: List[str] = Body(..., embed=True)):
    result = await db.stock_movements.delete_many({"id": {"$in": ids}})
    return {"success": True, "deleted": result.deleted_count}


class TransferMagasinCuisineRequest(BaseModel):
    source_product_id: str  # product in "magasin" zone
    target_product_id: Optional[str] = None  # optional: existing cuisine product
    target_name: Optional[str] = None  # OR create a new cuisine product with this name
    target_category_id: Optional[str] = None  # only used when creating
    target_unit: Optional[str] = None  # only used when creating
    quantity: float
    reason: Optional[str] = None
    user_name: Optional[str] = "Administrateur"


@router.post("/transfer-magasin-cuisine")
async def transfer_magasin_to_cuisine(req: TransferMagasinCuisineRequest):
    """Atomic transfer from a magasin product to a cuisine product.

    - Decrements the magasin product by `quantity` (creates a 'transfert_sortie' movement).
    - Increments (or creates) the cuisine product by the same quantity (creates a 'transfert_entree' movement).
    - The two movements are linked via a shared `transfer_id`.
    """
    import uuid as _uuid

    # 1. Validate source
    source = await db.stock_products.find_one({"id": req.source_product_id})
    if not source:
        raise HTTPException(404, "Produit magasin source introuvable")
    if source.get("storage_zone") != "magasin":
        raise HTTPException(400, "Le produit source n'est pas en zone magasin")
    if req.quantity <= 0:
        raise HTTPException(400, "La quantité doit être strictement positive")
    src_qty = source.get("quantity", 0) or 0
    if req.quantity > src_qty:
        raise HTTPException(400, f"Stock magasin insuffisant. Disponible: {src_qty} {source.get('unit','')}")

    # 2. Resolve or create target (cuisine product)
    target = None
    if req.target_product_id:
        target = await db.stock_products.find_one({"id": req.target_product_id})
        if not target:
            raise HTTPException(404, "Produit cuisine cible introuvable")
        if target.get("storage_zone") == "magasin":
            raise HTTPException(400, "Le produit cible est en magasin — choisissez un produit cuisine")
    else:
        target_name = (req.target_name or source.get("name") or "").strip()
        if not target_name:
            raise HTTPException(400, "Nom du produit cuisine cible manquant")
        existing = await db.stock_products.find_one({
            "name": target_name,
            "storage_zone": {"$ne": "magasin"}
        })
        if existing:
            target = existing
        else:
            target = {
                "id": str(_uuid.uuid4()),
                "code": f"PRD-{str(_uuid.uuid4())[:6].upper()}",
                "name": target_name,
                "category_id": req.target_category_id or source.get("category_id", ""),
                "subcategory": source.get("subcategory", ""),
                "unit": req.target_unit or source.get("unit", "unit"),
                "quantity": 0,
                "stock_min": source.get("stock_min", 0),
                "stock_max": source.get("stock_max", 0),
                "purchase_price": source.get("purchase_price", 0),
                "sale_price": source.get("sale_price", 0),
                "valeur_stock": 0,
                "valeur_stock_vente": 0,
                "supplier_id": source.get("supplier_id", ""),
                "storage_location": source.get("storage_location", ""),
                "storage_zone": "cuisine",
                "is_active": True,
                "photo_url": source.get("photo_url", ""),
                "statut": "rupture",
                "created_at": datetime.now(timezone.utc).isoformat(),
                "updated_at": datetime.now(timezone.utc).isoformat(),
            }
            await db.stock_products.insert_one(target.copy())
            target.pop("_id", None)

    # 3. Compute new quantities
    transfer_id = str(_uuid.uuid4())
    src_new = src_qty - req.quantity
    tgt_qty = target.get("quantity", 0) or 0
    tgt_new = tgt_qty + req.quantity

    reason = req.reason or f"Transfert magasin → cuisine ({target['name']})"
    now_iso = datetime.now(timezone.utc).isoformat()

    mov_out = {
        "id": str(_uuid.uuid4()),
        "product_id": source["id"],
        "product_name": source.get("name", ""),
        "product_code": source.get("code", ""),
        "movement_type": "transfert_sortie",
        "quantity": req.quantity,
        "previous_quantity": src_qty,
        "new_quantity": src_new,
        "unit": source.get("unit", ""),
        "unit_price": source.get("purchase_price", 0),
        "total_value": req.quantity * (source.get("purchase_price", 0) or 0),
        "reason": reason,
        "user_name": req.user_name or "Administrateur",
        "transfer_id": transfer_id,
        "transfer_role": "source",
        "linked_product_id": target["id"],
        "created_at": now_iso,
    }
    mov_in = {
        "id": str(_uuid.uuid4()),
        "product_id": target["id"],
        "product_name": target.get("name", ""),
        "product_code": target.get("code", ""),
        "movement_type": "transfert_entree",
        "quantity": req.quantity,
        "previous_quantity": tgt_qty,
        "new_quantity": tgt_new,
        "unit": target.get("unit", ""),
        "unit_price": source.get("purchase_price", 0),
        "total_value": req.quantity * (source.get("purchase_price", 0) or 0),
        "reason": f"Reçu de magasin ({source['name']})",
        "user_name": req.user_name or "Administrateur",
        "transfer_id": transfer_id,
        "transfer_role": "target",
        "linked_product_id": source["id"],
        "created_at": now_iso,
    }

    await db.stock_movements.insert_many([mov_out.copy(), mov_in.copy()])
    await db.stock_products.update_one(
        {"id": source["id"]},
        {"$set": {
            "quantity": src_new,
            "valeur_stock": src_new * (source.get("purchase_price", 0) or 0),
            "statut": "rupture" if src_new <= 0 else ("faible" if src_new <= source.get("stock_min", 0) else "normal"),
            "updated_at": now_iso,
        }},
    )
    await db.stock_products.update_one(
        {"id": target["id"]},
        {"$set": {
            "quantity": tgt_new,
            "valeur_stock": tgt_new * (target.get("purchase_price", 0) or 0),
            "statut": "rupture" if tgt_new <= 0 else ("faible" if tgt_new <= target.get("stock_min", 0) else "normal"),
            "updated_at": now_iso,
        }},
    )

    mov_out.pop("_id", None)
    mov_in.pop("_id", None)
    return {
        "success": True,
        "transfer_id": transfer_id,
        "source_movement": mov_out,
        "target_movement": mov_in,
        "target_product_id": target["id"],
    }



class ConvertUnitRequest(BaseModel):
    multiplier: int  # e.g. 24 bouteilles per casier
    new_unit: str    # e.g. "bouteille"


@router.post("/products/{product_id}/convert-unit")
async def convert_product_unit(product_id: str, data: ConvertUnitRequest):
    """Convert a product from a package unit (casier, pack, ...) to its inner unit
    (bouteille, ...). Multiplies quantity and stock_min by `multiplier`, divides
    purchase_price by `multiplier`, and updates the unit label. The total value
    (qty * price) is preserved."""
    if data.multiplier <= 0:
        raise HTTPException(400, "Le multiplicateur doit être > 0")
    if not data.new_unit.strip():
        raise HTTPException(400, "La nouvelle unité est requise")

    product = await db.stock_products.find_one({"id": product_id})
    if not product:
        raise HTTPException(404, "Produit non trouve")

    old_qty = product.get("quantity", 0) or 0
    old_price = product.get("purchase_price", 0) or 0
    old_min = product.get("stock_min", 0) or 0
    old_max = product.get("stock_max", 0) or 0
    old_unit = product.get("unit", "")

    new_qty = old_qty * data.multiplier
    new_price = old_price / data.multiplier if old_price else 0.0
    new_min = old_min * data.multiplier
    new_max = old_max * data.multiplier if old_max else 0
    new_valeur = new_qty * new_price
    smin = new_min or 5
    new_statut = "rupture" if new_qty <= 0 else ("faible" if new_qty <= smin else "normal")

    now_iso = datetime.now(timezone.utc).isoformat()
    await db.stock_products.update_one(
        {"id": product_id},
        {"$set": {
            "quantity": new_qty,
            "purchase_price": new_price,
            "stock_min": new_min,
            "stock_max": new_max,
            "unit": data.new_unit.strip(),
            "valeur_stock": new_valeur,
            "statut": new_statut,
            "updated_at": now_iso,
            "observation": (product.get("observation") or "") + f" [Unité convertie: {old_unit}×{data.multiplier} → {data.new_unit.strip()} le {now_iso[:10]}]",
        }}
    )

    # Trace the conversion as a special movement
    await db.stock_movements.insert_one({
        "id": str(uuid.uuid4()),
        "product_id": product_id,
        "product_name": product.get("name", ""),
        "product_code": product.get("code", ""),
        "movement_type": "conversion",
        "quantity": new_qty - old_qty,
        "previous_quantity": old_qty,
        "new_quantity": new_qty,
        "unit": data.new_unit.strip(),
        "unit_price": new_price,
        "total_value": 0,  # value unchanged
        "reason": f"Conversion {old_unit} → {data.new_unit.strip()} (×{data.multiplier})",
        "user_name": "Admin",
        "created_at": now_iso,
    })

    updated = await db.stock_products.find_one({"id": product_id}, {"_id": 0})
    return {"success": True, "product": updated}


class AddPackageRequest(BaseModel):
    package_qty: float           # ex: 2 (casiers achetés)
    package_price: float         # ex: 7200 F par casier
    items_per_package: int       # ex: 24 bouteilles par casier
    reason: str = ""             # optional note


@router.post("/products/{product_id}/add-package")
async def add_package_entry(product_id: str, data: AddPackageRequest):
    """Ajoute une entrée en stock par package (casier/pack/...) : multiplie la qty
    par items_per_package, divise le prix pour obtenir le PU à l'unité interne.
    Met à jour le prix d'achat du produit avec le nouveau PU calculé. Trace un
    stock_movements entree standard pour audit."""
    if data.package_qty <= 0:
        raise HTTPException(400, "Nombre de packages > 0 requis")
    if data.items_per_package <= 0:
        raise HTTPException(400, "Nombre d'unités par package > 0 requis")
    if data.package_price < 0:
        raise HTTPException(400, "Prix invalide")

    product = await db.stock_products.find_one({"id": product_id})
    if not product:
        raise HTTPException(404, "Produit non trouvé")

    n_units = data.package_qty * data.items_per_package
    unit_price = data.package_price / data.items_per_package if data.items_per_package else 0.0

    old_qty = product.get("quantity", 0) or 0
    new_qty = old_qty + n_units
    new_valeur = new_qty * unit_price
    smin = product.get("stock_min", 5)
    new_statut = "rupture" if new_qty <= 0 else ("faible" if new_qty <= smin else "normal")

    now_iso = datetime.now(timezone.utc).isoformat()
    await db.stock_products.update_one(
        {"id": product_id},
        {"$set": {
            "quantity": new_qty,
            "purchase_price": unit_price,
            "valeur_stock": new_valeur,
            "statut": new_statut,
            "updated_at": now_iso,
            "date_achat": now_iso[:10],
        }}
    )

    await db.stock_movements.insert_one({
        "id": str(uuid.uuid4()),
        "product_id": product_id,
        "product_name": product.get("name", ""),
        "product_code": product.get("code", ""),
        "movement_type": "entree",
        "quantity": n_units,
        "previous_quantity": old_qty,
        "new_quantity": new_qty,
        "unit": product.get("unit", ""),
        "unit_price": unit_price,
        "total_value": data.package_qty * data.package_price,
        "reason": data.reason or f"Entrée par package ({data.package_qty} × {data.items_per_package} @ {data.package_price} F)",
        "user_name": "Admin",
        "created_at": now_iso,
    })

    updated = await db.stock_products.find_one({"id": product_id}, {"_id": 0})
    return {"success": True, "product": updated, "units_added": n_units, "new_unit_price": unit_price}


class ConvertUnitBulkRequest(BaseModel):
    category_id: Optional[str] = None  # optional filter by category
    from_unit: str                      # e.g. "casier" — only convert products with this current unit
    multiplier: int                     # e.g. 24
    new_unit: str                       # e.g. "bouteille"


@router.post("/products/convert-unit-bulk")
async def convert_products_unit_bulk(data: ConvertUnitBulkRequest):
    """Bulk conversion: apply unit conversion to every active product matching
    (category_id, from_unit). Each product's quantity × multiplier, price ÷ multiplier.
    The total value is preserved for each product."""
    if data.multiplier <= 0:
        raise HTTPException(400, "Le multiplicateur doit être > 0")
    if not data.new_unit.strip():
        raise HTTPException(400, "La nouvelle unité est requise")
    if not data.from_unit.strip():
        raise HTTPException(400, "L'unité actuelle (from_unit) est requise")

    query = {"is_active": True, "unit": {"$regex": f"^{re.escape(data.from_unit)}$", "$options": "i"}}
    if data.category_id:
        query["category_id"] = data.category_id

    products = await db.stock_products.find(query).to_list(1000)
    if not products:
        return {"success": True, "converted": 0, "products": []}

    now_iso = datetime.now(timezone.utc).isoformat()
    converted = []
    for product in products:
        old_qty = product.get("quantity", 0) or 0
        old_price = product.get("purchase_price", 0) or 0
        old_min = product.get("stock_min", 0) or 0
        old_max = product.get("stock_max", 0) or 0
        old_unit = product.get("unit", "")

        new_qty = old_qty * data.multiplier
        new_price = old_price / data.multiplier if old_price else 0.0
        new_min = old_min * data.multiplier
        new_max = old_max * data.multiplier if old_max else 0
        new_valeur = new_qty * new_price
        smin = new_min or 5
        new_statut = "rupture" if new_qty <= 0 else ("faible" if new_qty <= smin else "normal")

        await db.stock_products.update_one(
            {"id": product["id"]},
            {"$set": {
                "quantity": new_qty,
                "purchase_price": new_price,
                "stock_min": new_min,
                "stock_max": new_max,
                "unit": data.new_unit.strip(),
                "valeur_stock": new_valeur,
                "statut": new_statut,
                "updated_at": now_iso,
                "observation": (product.get("observation") or "") + f" [Unité convertie en lot: {old_unit}×{data.multiplier} → {data.new_unit.strip()} le {now_iso[:10]}]",
            }}
        )
        await db.stock_movements.insert_one({
            "id": str(uuid.uuid4()),
            "product_id": product["id"],
            "product_name": product.get("name", ""),
            "product_code": product.get("code", ""),
            "movement_type": "conversion",
            "quantity": new_qty - old_qty,
            "previous_quantity": old_qty,
            "new_quantity": new_qty,
            "unit": data.new_unit.strip(),
            "unit_price": new_price,
            "total_value": 0,
            "reason": f"Conversion en lot {old_unit} → {data.new_unit.strip()} (×{data.multiplier})",
            "user_name": "Admin",
            "created_at": now_iso,
        })
        converted.append({"id": product["id"], "name": product.get("name"), "old_qty": old_qty, "new_qty": new_qty})

    return {"success": True, "converted": len(converted), "products": converted}

# ==================== SUPPLIERS ====================

@router.get("/suppliers")
async def get_suppliers():
    suppliers = await db.stock_suppliers.find({}, {"_id": 0}).sort("name", 1).to_list(500)
    return {"suppliers": suppliers}

@router.post("/suppliers")
async def create_supplier(data: StockSupplierCreate):
    supplier = {
        "id": str(uuid.uuid4()),
        "name": data.name,
        "phone": data.phone,
        "email": data.email,
        "address": data.address,
        "product_types": data.product_types,
        "notes": data.notes,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.stock_suppliers.insert_one(supplier)
    supplier.pop("_id", None)
    return {"success": True, "supplier": supplier}

@router.put("/suppliers/{supplier_id}")
async def update_supplier(supplier_id: str, data: dict = Body(...)):
    data.pop("_id", None)
    data.pop("id", None)
    await db.stock_suppliers.update_one({"id": supplier_id}, {"$set": data})
    updated = await db.stock_suppliers.find_one({"id": supplier_id}, {"_id": 0})
    return {"success": True, "supplier": updated}

@router.delete("/suppliers/{supplier_id}")
async def delete_supplier(supplier_id: str):
    await db.stock_suppliers.delete_one({"id": supplier_id})
    return {"success": True}

# ==================== PURCHASES ====================

@router.get("/purchases")
async def get_purchases(supplier_id: str = None, date_from: str = None, date_to: str = None):
    import re as _re
    query = {}
    if supplier_id:
        query["supplier_id"] = supplier_id
    if date_from or date_to:
        dq = {}
        if date_from: dq["$gte"] = date_from
        if date_to: dq["$lte"] = date_to + "T23:59:59"
        query["purchase_date"] = dq
    purchases = await db.stock_purchases.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    
    # Also include caisse expenses not yet synced to stock_purchases
    # BUT only items that match actual stock products
    synced_expense_ids = {p.get("expense_id") for p in purchases if p.get("expense_id")}
    
    expense_query = {}
    if date_from or date_to:
        eq = {}
        if date_from: eq["$gte"] = date_from
        if date_to: eq["$lte"] = date_to + "T23:59:59"
        expense_query["created_at"] = eq
    
    caisse_expenses = await db.expenses.find(expense_query, {"_id": 0}).sort("created_at", -1).to_list(500)
    
    # Load all stock product names for matching
    all_stock_names = [p["name"].lower() for p in await db.stock_products.find({"is_active": True}, {"name": 1, "_id": 0}).to_list(5000)]
    
    for exp in caisse_expenses:
        if exp.get("id") in synced_expense_ids:
            continue
        
        raw_items = []
        if exp.get("is_group") and exp.get("items"):
            raw_items = exp["items"]
        else:
            raw_items = [{"description": exp.get("description", ""), "quantity": exp.get("quantity", 1), "unit_price": exp.get("unit_price", 0) or exp.get("amount", 0)}]
        
        # Filter: only keep items that match a stock product (exact or starts-with match)
        matched_items = []
        for item in raw_items:
            desc = (item.get("description") or "").strip().lower()
            if not desc or len(desc) < 2:
                continue
            # Exact match or stock product name starts with the description
            found = any(desc == sn or sn.startswith(desc + " ") or sn.startswith(desc + "s") or desc.rstrip("s") == sn for sn in all_stock_names)
            if found:
                matched_items.append({
                    "product_name": item.get("description", ""),
                    "quantity": item.get("quantity", 1),
                    "unit_price": item.get("unit_price", 0)
                })
        
        if not matched_items:
            continue
        
        matched_total = sum(i["quantity"] * i["unit_price"] for i in matched_items)
        status_map = {"pending": "en_attente", "approved": "approuve", "revision_requested": "en_revision", "rejected": "rejete", "completed": "valide"}
        
        purchases.append({
            "id": f"caisse-{exp['id']}",
            "supplier_id": "",
            "supplier_name": exp.get("supplier") or exp.get("description", "Caisse"),
            "purchase_date": exp.get("created_at", "")[:10],
            "items": matched_items,
            "total_amount": matched_total,
            "notes": exp.get("description", ""),
            "user_name": exp.get("requested_by", ""),
            "status": status_map.get(exp.get("status", ""), exp.get("status", "")),
            "source": "caisse",
            "expense_id": exp.get("id"),
            "caisse_status": exp.get("status", ""),
            "created_at": exp.get("created_at", "")
        })
    
    purchases.sort(key=lambda x: x.get("created_at", ""), reverse=True)
    
    return {"purchases": purchases}

@router.post("/purchases")
async def create_purchase(data: StockPurchaseCreate):
    total = sum(item.get("quantity", 0) * item.get("unit_price", 0) for item in data.items)
    
    purchase = {
        "id": str(uuid.uuid4()),
        "supplier_id": data.supplier_id,
        "supplier_name": data.supplier_name,
        "purchase_date": data.purchase_date or datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        "items": data.items,
        "total_amount": total,
        "notes": data.notes,
        "user_name": data.user_name,
        "status": "validated",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.stock_purchases.insert_one(purchase)
    purchase.pop("_id", None)
    
    # Update stock for each item
    for item in data.items:
        pid = item.get("product_id")
        qty = item.get("quantity", 0)
        price = item.get("unit_price", 0)
        if pid and qty > 0:
            product = await db.stock_products.find_one({"id": pid})
            if product:
                old_qty = product.get("quantity", 0)
                new_qty = old_qty + qty
                await db.stock_products.update_one(
                    {"id": pid},
                    {"$set": {"quantity": new_qty, "purchase_price": price, "updated_at": datetime.now(timezone.utc).isoformat()}}
                )
                # Create movement
                mov = {
                    "id": str(uuid.uuid4()),
                    "product_id": pid,
                    "product_name": product.get("name", ""),
                    "product_code": product.get("code", ""),
                    "movement_type": "entree",
                    "quantity": qty,
                    "previous_quantity": old_qty,
                    "new_quantity": new_qty,
                    "unit": product.get("unit", ""),
                    "unit_price": price,
                    "total_value": qty * price,
                    "reason": f"Achat - {data.supplier_name}",
                    "user_name": data.user_name,
                    "purchase_id": purchase["id"],
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                await db.stock_movements.insert_one(mov)
    
    return {"success": True, "purchase": purchase}

@router.delete("/purchases/{purchase_id}")
async def delete_purchase(purchase_id: str):
    result = await db.stock_purchases.delete_one({"id": purchase_id})
    if result.deleted_count == 0:
        raise HTTPException(404, "Achat non trouve")
    return {"success": True}

@router.post("/purchases/delete-bulk")
async def delete_purchases_bulk(ids: List[str] = Body(..., embed=True)):
    result = await db.stock_purchases.delete_many({"id": {"$in": ids}})
    return {"success": True, "deleted": result.deleted_count}

# ==================== BULK DELETE (ALL SECTIONS) ====================

@router.post("/products/delete-bulk")
async def delete_products_bulk(ids: List[str] = Body(..., embed=True)):
    result = await db.stock_products.delete_many({"id": {"$in": ids}})
    return {"success": True, "deleted": result.deleted_count}

@router.post("/suppliers/delete-bulk")
async def delete_suppliers_bulk(ids: List[str] = Body(..., embed=True)):
    result = await db.stock_suppliers.delete_many({"id": {"$in": ids}})
    return {"success": True, "deleted": result.deleted_count}

@router.post("/categories/delete-bulk")
async def delete_categories_bulk(ids: List[str] = Body(..., embed=True)):
    result = await db.stock_categories.delete_many({"id": {"$in": ids}})
    return {"success": True, "deleted": result.deleted_count}

@router.post("/recipes/delete-bulk")
async def delete_recipes_bulk(ids: List[str] = Body(..., embed=True)):
    result = await db.stock_recipes.delete_many({"id": {"$in": ids}})
    return {"success": True, "deleted": result.deleted_count}

@router.post("/auth/users/delete-bulk")
async def delete_users_bulk(ids: List[str] = Body(..., embed=True)):
    result = await db.stock_users.delete_many({"id": {"$in": ids}})
    return {"success": True, "deleted": result.deleted_count}

@router.post("/products/reset-quantities")
async def reset_product_quantities(ids: List[str] = Body(..., embed=True)):
    """Reset stock quantities to 0 for given product IDs"""
    now_iso = datetime.now(timezone.utc).isoformat()
    reset_count = 0
    for pid in ids:
        product = await db.stock_products.find_one({"id": pid})
        if not product:
            continue
        old_qty = product.get("quantity", 0)
        if old_qty == 0:
            continue
        # Create movement for traceability
        mov = {
            "id": str(uuid.uuid4()),
            "product_id": pid,
            "product_name": product.get("name", ""),
            "product_code": product.get("code", ""),
            "movement_type": "ajustement",
            "quantity": old_qty,
            "previous_quantity": old_qty,
            "new_quantity": 0,
            "unit": product.get("unit", ""),
            "unit_price": product.get("purchase_price", 0),
            "total_value": round(old_qty * product.get("purchase_price", 0), 2),
            "reason": "Reinitialisation a zero",
            "user_name": "Admin",
            "created_at": now_iso
        }
        await db.stock_movements.insert_one(mov)
        await db.stock_products.update_one(
            {"id": pid},
            {"$set": {"quantity": 0, "valeur_stock": 0, "statut": "rupture", "updated_at": now_iso}}
        )
        reset_count += 1
    return {"success": True, "reset": reset_count}

@router.post("/products/reset-prices")
async def reset_product_prices(ids: List[str] = Body(..., embed=True)):
    """Reset purchase prices to 0 for given product IDs"""
    now_iso = datetime.now(timezone.utc).isoformat()
    reset_count = 0
    for pid in ids:
        product = await db.stock_products.find_one({"id": pid})
        if not product or product.get("purchase_price", 0) == 0:
            continue
        qty = product.get("quantity", 0)
        await db.stock_products.update_one(
            {"id": pid},
            {"$set": {"purchase_price": 0, "valeur_stock": 0, "updated_at": now_iso}}
        )
        reset_count += 1
    return {"success": True, "reset": reset_count}

# ==================== INVENTAIRE PHYSIQUE ====================

class InventoryItemInput(BaseModel):
    product_id: str
    physical_quantity: float

class InventoryCreate(BaseModel):
    name: str = ""
    notes: str = ""
    category_id: str = ""  # empty = all categories
    user_name: str = ""

@router.post("/inventories")
async def create_inventory(data: InventoryCreate):
    """Create a new inventory session with all active products (or filtered by category)"""
    query = {"is_active": True}
    if data.category_id:
        query["category_id"] = data.category_id
    
    products = await db.stock_products.find(query, {"_id": 0}).sort("name", 1).to_list(5000)
    
    items = []
    for p in products:
        items.append({
            "product_id": p["id"],
            "product_name": p.get("name", ""),
            "product_code": p.get("code", ""),
            "category_id": p.get("category_id", ""),
            "unit": p.get("unit", ""),
            "theoretical_quantity": p.get("quantity", 0),
            "physical_quantity": None,  # Not yet counted
            "ecart": 0,
            "ecart_value": 0,
            "purchase_price": p.get("purchase_price", 0),
            "counted": False
        })
    
    inventory = {
        "id": str(uuid.uuid4()),
        "name": data.name or f"Inventaire du {datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%M')}",
        "status": "en_cours",  # en_cours, valide, annule
        "category_id": data.category_id,
        "items": items,
        "total_products": len(items),
        "counted_products": 0,
        "total_ecart_value": 0,
        "notes": data.notes,
        "created_by": data.user_name,
        "validated_by": "",
        "validated_at": "",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.stock_inventories.insert_one(inventory)
    inventory.pop("_id", None)
    return {"success": True, "inventory": inventory}

@router.get("/inventories")
async def get_inventories():
    inventories = await db.stock_inventories.find({}, {"_id": 0, "items": 0}).sort("created_at", -1).to_list(100)
    return {"inventories": inventories}

@router.get("/inventories/{inventory_id}")
async def get_inventory(inventory_id: str):
    inv = await db.stock_inventories.find_one({"id": inventory_id}, {"_id": 0})
    if not inv:
        raise HTTPException(404, "Inventaire non trouve")
    
    # Refresh theoretical quantities from current stock
    if inv.get("status") == "en_cours":
        for item in inv.get("items", []):
            product = await db.stock_products.find_one({"id": item["product_id"]}, {"_id": 0})
            if product:
                item["theoretical_quantity"] = product.get("quantity", 0)
                item["purchase_price"] = product.get("purchase_price", 0)
                if item.get("counted") and item.get("physical_quantity") is not None:
                    item["ecart"] = round(item["physical_quantity"] - item["theoretical_quantity"], 3)
                    item["ecart_value"] = round(item["ecart"] * item["purchase_price"], 2)
    
    # Recalc stats
    counted = sum(1 for i in inv.get("items", []) if i.get("counted"))
    total_ecart = sum(i.get("ecart_value", 0) for i in inv.get("items", []) if i.get("counted"))
    inv["counted_products"] = counted
    inv["total_ecart_value"] = round(total_ecart, 2)
    
    return {"inventory": inv}

@router.put("/inventories/{inventory_id}/count")
async def update_inventory_count(inventory_id: str, items: List[InventoryItemInput] = Body(..., embed=True)):
    """Update physical counts for products in an inventory"""
    inv = await db.stock_inventories.find_one({"id": inventory_id})
    if not inv:
        raise HTTPException(404, "Inventaire non trouve")
    if inv.get("status") != "en_cours":
        raise HTTPException(400, "Cet inventaire est deja cloture")
    
    existing_items = inv.get("items", [])
    item_map = {i["product_id"]: i for i in existing_items}
    
    for update in items:
        if update.product_id in item_map:
            item = item_map[update.product_id]
            item["physical_quantity"] = update.physical_quantity
            item["counted"] = True
            item["ecart"] = round(update.physical_quantity - item["theoretical_quantity"], 3)
            item["ecart_value"] = round(item["ecart"] * item.get("purchase_price", 0), 2)
    
    counted = sum(1 for i in existing_items if i.get("counted"))
    total_ecart = sum(i.get("ecart_value", 0) for i in existing_items if i.get("counted"))
    
    await db.stock_inventories.update_one(
        {"id": inventory_id},
        {"$set": {
            "items": existing_items,
            "counted_products": counted,
            "total_ecart_value": round(total_ecart, 2),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"success": True, "counted_products": counted, "total_ecart_value": round(total_ecart, 2)}

@router.put("/inventories/{inventory_id}/validate")
async def validate_inventory(inventory_id: str, user_name: str = Body("", embed=True)):
    """Validate inventory and adjust stock quantities to match physical counts"""
    inv = await db.stock_inventories.find_one({"id": inventory_id})
    if not inv:
        raise HTTPException(404, "Inventaire non trouve")
    if inv.get("status") != "en_cours":
        raise HTTPException(400, "Cet inventaire est deja cloture")
    
    now_iso = datetime.now(timezone.utc).isoformat()
    adjustments = 0
    
    for item in inv.get("items", []):
        if not item.get("counted") or item.get("physical_quantity") is None:
            continue
        
        phys_qty = item["physical_quantity"]
        theo_qty = item["theoretical_quantity"]
        ecart = round(phys_qty - theo_qty, 3)
        
        if ecart == 0:
            continue
        
        # Create adjustment movement
        product = await db.stock_products.find_one({"id": item["product_id"]})
        if not product:
            continue
        
        current_qty = product.get("quantity", 0)
        new_qty = phys_qty
        price = product.get("purchase_price", 0)
        new_valeur = round(new_qty * price, 2)
        smin = product.get("stock_min", 5)
        new_statut = "rupture" if new_qty <= 0 else ("faible" if new_qty <= smin else "normal")
        
        mov = {
            "id": str(uuid.uuid4()),
            "product_id": item["product_id"],
            "product_name": item["product_name"],
            "product_code": item.get("product_code", ""),
            "movement_type": "inventaire",
            "quantity": abs(ecart),
            "previous_quantity": current_qty,
            "new_quantity": new_qty,
            "unit": item.get("unit", ""),
            "unit_price": price,
            "total_value": round(abs(ecart) * price, 2),
            "reason": f"Ajustement inventaire - {inv.get('name', '')}",
            "user_name": user_name or inv.get("created_by", ""),
            "inventory_id": inventory_id,
            "created_at": now_iso
        }
        await db.stock_movements.insert_one(mov)
        
        await db.stock_products.update_one(
            {"id": item["product_id"]},
            {"$set": {
                "quantity": new_qty,
                "valeur_stock": new_valeur,
                "statut": new_statut,
                "updated_at": now_iso
            }}
        )
        adjustments += 1
    
    await db.stock_inventories.update_one(
        {"id": inventory_id},
        {"$set": {
            "status": "valide",
            "validated_by": user_name,
            "validated_at": now_iso,
            "updated_at": now_iso
        }}
    )
    
    return {"success": True, "adjustments": adjustments, "message": f"Inventaire valide. {adjustments} produit(s) ajuste(s)."}

@router.delete("/inventories/{inventory_id}")
async def delete_inventory(inventory_id: str):
    result = await db.stock_inventories.delete_one({"id": inventory_id})
    if result.deleted_count == 0:
        raise HTTPException(404, "Inventaire non trouve")
    return {"success": True}

# ==================== DASHBOARD ====================

@router.get("/dashboard")
async def get_dashboard():
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    peremption_limit = (datetime.now(timezone.utc) + timedelta(days=7)).strftime("%Y-%m-%d")
    
    products = await db.stock_products.find({"is_active": True}, {"_id": 0}).to_list(5000)
    categories = await db.stock_categories.find({}, {"_id": 0}).to_list(100)
    
    total_products = len(products)
    total_value = sum(p.get("quantity", 0) * p.get("purchase_price", 0) for p in products)
    total_value_vente = sum(p.get("quantity", 0) * p.get("sale_price", 0) for p in products)
    
    rupture = [p for p in products if p.get("quantity", 0) <= 0]
    faible = [p for p in products if 0 < p.get("quantity", 0) <= p.get("stock_min", 5)]
    critical = len(rupture) + len(faible)
    
    # Products near expiry
    peremption_proche = [p for p in products if p.get("date_peremption") and p["date_peremption"] <= peremption_limit and p["date_peremption"] >= today]
    expired = [p for p in products if p.get("date_peremption") and p["date_peremption"] < today]
    
    # Today's movements
    today_movements = await db.stock_movements.find(
        {"created_at": {"$gte": today}}, {"_id": 0}
    ).sort("created_at", -1).to_list(500)
    
    entrees_today = sum(1 for m in today_movements if m.get("movement_type") in ["entree", "retour_fournisseur"])
    sorties_today = sum(1 for m in today_movements if m.get("movement_type") in ["sortie", "perte", "casse"])
    
    recent_movements = await db.stock_movements.find({}, {"_id": 0}).sort("created_at", -1).to_list(10)
    recent_purchases = await db.stock_purchases.find({}, {"_id": 0}).sort("created_at", -1).to_list(5)
    
    # Stock by category
    cat_map = {c["id"]: c["name"] for c in categories}
    stock_by_category = {}
    for p in products:
        cname = cat_map.get(p.get("category_id"), "Autre")
        if cname not in stock_by_category:
            stock_by_category[cname] = {"count": 0, "value": 0, "value_vente": 0}
        stock_by_category[cname]["count"] += 1
        stock_by_category[cname]["value"] += p.get("quantity", 0) * p.get("purchase_price", 0)
        stock_by_category[cname]["value_vente"] += p.get("quantity", 0) * p.get("sale_price", 0)
    
    # Top sorted products (most movements)
    top_sorted = await db.stock_movements.aggregate([
        {"$match": {"movement_type": "sortie"}},
        {"$group": {"_id": "$product_name", "total": {"$sum": "$quantity"}}},
        {"$sort": {"total": -1}},
        {"$limit": 10}
    ]).to_list(10)
    
    # Today's sales from Caisse
    today_sales = await db.stock_movements.find(
        {"created_at": {"$gte": today}, "invoice_id": {"$exists": True}}, {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    ventes_caisse_today = len(today_sales)
    ventes_caisse_value = sum(m.get("total_value", 0) for m in today_sales)
    
    return {
        "total_products": total_products,
        "critical_products": critical,
        "total_value": total_value,
        "total_value_vente": total_value_vente,
        "potential_margin": total_value_vente - total_value,
        "entrees_today": entrees_today,
        "sorties_today": sorties_today,
        "ventes_caisse_today": ventes_caisse_today,
        "ventes_caisse_value": ventes_caisse_value,
        "rupture": [{"id": p["id"], "name": p["name"], "code": p.get("code", ""), "unit": p.get("unit", "")} for p in rupture[:20]],
        "faible": [{"id": p["id"], "name": p["name"], "quantity": p["quantity"], "stock_min": p.get("stock_min", 5), "unit": p.get("unit", "")} for p in faible[:20]],
        "peremption_proche": [{"id": p["id"], "name": p["name"], "date_peremption": p["date_peremption"]} for p in peremption_proche[:10]],
        "expired": [{"id": p["id"], "name": p["name"], "date_peremption": p["date_peremption"]} for p in expired[:10]],
        "recent_movements": recent_movements,
        "recent_purchases": recent_purchases,
        "stock_by_category": stock_by_category,
        "top_sorted": top_sorted
    }

# ==================== RECIPES / FICHES TECHNIQUES ====================

class RecipeIngredient(BaseModel):
    product_id: str
    product_name: str = ""
    quantity: float
    unit: str = ""

class StockRecipeCreate(BaseModel):
    name: str
    caisse_product_name: str  # Name as it appears on Caisse invoices
    selling_price: float = 0
    ingredients: List[RecipeIngredient] = []
    notes: str = ""

class StockRecipeUpdate(BaseModel):
    name: Optional[str] = None
    caisse_product_name: Optional[str] = None
    selling_price: Optional[float] = None
    ingredients: Optional[List[RecipeIngredient]] = None
    notes: Optional[str] = None

@router.get("/recipes")
async def get_recipes():
    recipes = await db.stock_recipes.find({}, {"_id": 0}).sort("name", 1).to_list(500)
    # Enrich with current cost prices from products
    all_products = {p["id"]: p for p in await db.stock_products.find({"is_active": True}, {"_id": 0}).to_list(5000)}
    for r in recipes:
        cost = 0
        for ing in r.get("ingredients", []):
            prod = all_products.get(ing.get("product_id"))
            if prod:
                ing["current_stock"] = prod.get("quantity", 0)
                ing["purchase_price"] = prod.get("purchase_price", 0)
                ing["unit"] = ing.get("unit") or prod.get("unit", "")
                cost += ing["quantity"] * prod.get("purchase_price", 0)
            else:
                ing["current_stock"] = 0
                ing["purchase_price"] = 0
        r["cost_price"] = round(cost, 2)
        r["margin"] = round(r.get("selling_price", 0) - cost, 2) if r.get("selling_price") else 0
        r["margin_percent"] = round((r["margin"] / r["selling_price"]) * 100, 1) if r.get("selling_price") and r["selling_price"] > 0 else 0
    return {"recipes": recipes}

@router.post("/recipes")
async def create_recipe(data: StockRecipeCreate):
    existing = await db.stock_recipes.find_one({"caisse_product_name": {"$regex": f"^{data.caisse_product_name}$", "$options": "i"}})
    if existing:
        raise HTTPException(400, f"Une fiche technique existe deja pour '{data.caisse_product_name}'")
    
    recipe = {
        "id": str(uuid.uuid4()),
        "name": data.name,
        "caisse_product_name": data.caisse_product_name,
        "selling_price": data.selling_price,
        "ingredients": [ing.dict() for ing in data.ingredients],
        "notes": data.notes,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    await db.stock_recipes.insert_one(recipe)
    recipe.pop("_id", None)
    return {"success": True, "recipe": recipe}

@router.put("/recipes/{recipe_id}")
async def update_recipe(recipe_id: str, data: StockRecipeUpdate):
    recipe = await db.stock_recipes.find_one({"id": recipe_id})
    if not recipe:
        raise HTTPException(404, "Fiche technique non trouvee")
    
    update_data = {}
    for k, v in data.dict().items():
        if v is not None:
            if k == "ingredients":
                update_data[k] = [ing if isinstance(ing, dict) else ing.dict() for ing in v]
            else:
                update_data[k] = v
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.stock_recipes.update_one({"id": recipe_id}, {"$set": update_data})
    updated = await db.stock_recipes.find_one({"id": recipe_id}, {"_id": 0})
    return {"success": True, "recipe": updated}

@router.delete("/recipes/{recipe_id}")
async def delete_recipe(recipe_id: str):
    result = await db.stock_recipes.delete_one({"id": recipe_id})
    if result.deleted_count == 0:
        raise HTTPException(404, "Fiche technique non trouvee")
    return {"success": True}

@router.post("/recipes/seed-demo")
async def seed_demo_recipes():
    """Seed demo recipes for Poulet braise"""
    existing = await db.stock_recipes.count_documents({})
    if existing > 0:
        return {"success": True, "message": f"{existing} fiche(s) technique(s) deja presente(s)"}
    
    # Find product IDs for Poulet braisé ingredients
    product_map = {}
    needed = ["Cuisses de poulet", "Oignon local", "Piment frais", "Huile d'arachide", 
              "Concentre de tomate", "Ail local", "Poivron vert", "Sel fin"]
    for name in needed:
        p = await db.stock_products.find_one({"name": name, "is_active": True})
        if p:
            product_map[name] = {"id": p["id"], "unit": p.get("unit", "")}
    
    recipes = []
    if product_map.get("Cuisses de poulet"):
        recipes.append({
            "id": str(uuid.uuid4()),
            "name": "Poulet braise",
            "caisse_product_name": "Poulet braise",
            "selling_price": 3500,
            "ingredients": [
                {"product_id": product_map["Cuisses de poulet"]["id"], "product_name": "Cuisses de poulet", "quantity": 0.5, "unit": "kg"},
                {"product_id": product_map.get("Oignon local", {}).get("id", ""), "product_name": "Oignon local", "quantity": 0.15, "unit": "kg"},
                {"product_id": product_map.get("Piment frais", {}).get("id", ""), "product_name": "Piment frais", "quantity": 0.05, "unit": "kg"},
                {"product_id": product_map.get("Huile d'arachide", {}).get("id", ""), "product_name": "Huile d'arachide", "quantity": 0.1, "unit": "litre"},
                {"product_id": product_map.get("Concentre de tomate", {}).get("id", ""), "product_name": "Concentre de tomate", "quantity": 0.05, "unit": "boite"},
                {"product_id": product_map.get("Ail local", {}).get("id", ""), "product_name": "Ail local", "quantity": 0.02, "unit": "kg"},
                {"product_id": product_map.get("Poivron vert", {}).get("id", ""), "product_name": "Poivron vert", "quantity": 0.1, "unit": "kg"},
                {"product_id": product_map.get("Sel fin", {}).get("id", ""), "product_name": "Sel fin", "quantity": 0.01, "unit": "kg"},
            ],
            "notes": "Recette standard pour 1 portion de poulet braise avec accompagnement",
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        })
        # Remove ingredients with empty product_id
        recipes[0]["ingredients"] = [i for i in recipes[0]["ingredients"] if i["product_id"]]
    
    if recipes:
        await db.stock_recipes.insert_many(recipes)
        # Clean _id
        for r in recipes:
            r.pop("_id", None)
    
    return {"success": True, "message": f"{len(recipes)} fiche(s) technique(s) de demo creee(s)", "recipes": recipes}


# ==================== AUTO-COMPOSE RECIPES ====================
# Smart matching of dish names → ingredients, based on keyword recognition.
# The portion sizes are conservative defaults intended for adjustment by the manager.

# Common cooking keywords (lowercase, accent-stripped) → ingredient categories.
# Each entry returns ingredient name patterns + a default portion in stock unit.
_DISH_KEYWORD_RULES = [
    # (keyword in dish name) -> [(stock product name regex, qty, unit_hint), ...]
    # Proteins
    ("poulet", [(r"poulet", 0.25, "kg"), (r"oignon", 0.05, "kg"), (r"tomate", 0.04, "kg"), (r"huile", 0.04, "litre"), (r"sel", 0.005, "kg")]),
    ("poisson", [(r"poisson|tilapia|carpe|capitaine|maquereau", 0.3, "kg"), (r"oignon", 0.05, "kg"), (r"tomate", 0.05, "kg"), (r"huile", 0.04, "litre"), (r"piment|epice", 0.01, "kg"), (r"sel", 0.005, "kg")]),
    ("viande", [(r"viande|boeuf|bovin", 0.2, "kg"), (r"oignon", 0.05, "kg"), (r"tomate", 0.04, "kg"), (r"huile", 0.04, "litre"), (r"sel", 0.005, "kg")]),
    ("boeuf", [(r"viande|boeuf", 0.2, "kg"), (r"oignon", 0.05, "kg"), (r"tomate", 0.04, "kg"), (r"huile", 0.04, "litre")]),
    ("agneau", [(r"agneau|mouton", 0.2, "kg"), (r"oignon", 0.05, "kg"), (r"epice|piment", 0.01, "kg")]),
    ("crevette", [(r"crevette", 0.15, "kg"), (r"oignon", 0.04, "kg"), (r"tomate", 0.04, "kg"), (r"huile", 0.03, "litre")]),
    ("samossa", [(r"farine|samossa", 0.05, "kg"), (r"viande|poulet", 0.06, "kg"), (r"oignon", 0.02, "kg"), (r"epice|piment", 0.005, "kg")]),
    ("nem", [(r"nem|farine", 0.05, "kg"), (r"poulet|crevette|viande", 0.06, "kg"), (r"chou|carotte", 0.04, "kg")]),

    # Bases / accompagnements
    ("riz", [(r"^riz", 0.18, "kg"), (r"huile", 0.02, "litre"), (r"sel", 0.005, "kg")]),
    ("frite", [(r"frite|pomme.*terre", 0.15, "kg"), (r"huile", 0.05, "litre"), (r"sel", 0.005, "kg")]),
    ("pate", [(r"pate|spaghetti|macaroni", 0.12, "kg"), (r"tomate", 0.05, "kg"), (r"huile", 0.02, "litre")]),
    ("attieke", [(r"attieke", 0.18, "kg"), (r"huile", 0.02, "litre")]),
    ("foufou", [(r"foufou|igname|manioc", 0.2, "kg")]),
    ("igname", [(r"igname", 0.2, "kg"), (r"huile", 0.02, "litre")]),
    ("manioc", [(r"manioc|placali", 0.18, "kg")]),
    ("haricot", [(r"haricot", 0.15, "kg"), (r"huile", 0.02, "litre")]),

    # Salades / Entrées
    ("salade", [(r"salade|laitue", 0.1, "kg"), (r"tomate", 0.05, "kg"), (r"oignon", 0.03, "kg"), (r"huile|vinaigre", 0.02, "litre"), (r"sel", 0.002, "kg")]),
    ("crudite", [(r"carotte|tomate|laitue|concombre|chou", 0.12, "kg"), (r"huile|vinaigre", 0.02, "litre")]),
    ("avocat", [(r"avocat", 0.1, "kg"), (r"laitue|salade", 0.05, "kg"), (r"tomate", 0.04, "kg")]),
    ("thon", [(r"thon", 0.08, "boite"), (r"laitue|salade", 0.05, "kg"), (r"tomate", 0.04, "kg")]),
    ("cesar", [(r"laitue|salade", 0.1, "kg"), (r"poulet", 0.08, "kg"), (r"parmesan|fromage", 0.02, "kg"), (r"croutons", 0.02, "kg")]),

    # Sauces
    ("sauce", [(r"sauce|tomate", 0.06, "kg"), (r"oignon", 0.04, "kg"), (r"huile", 0.03, "litre"), (r"piment", 0.005, "kg")]),
    ("graine", [(r"graine|palme", 0.1, "kg"), (r"poisson|viande", 0.15, "kg"), (r"oignon", 0.04, "kg")]),
    ("arachide", [(r"arachide|pate.*arachide", 0.08, "kg"), (r"viande|poisson", 0.15, "kg"), (r"oignon", 0.04, "kg")]),
    ("gombo", [(r"gombo", 0.1, "kg"), (r"poisson|viande", 0.15, "kg"), (r"huile", 0.03, "litre")]),
    ("tomate", [(r"tomate", 0.1, "kg"), (r"oignon", 0.04, "kg"), (r"huile", 0.03, "litre")]),
    ("aubergine", [(r"aubergine", 0.12, "kg"), (r"oignon", 0.04, "kg"), (r"huile", 0.03, "litre")]),

    # Boissons (contenant simple = 1 unité)
    ("biere", [(r"biere|castel|flag|beaufort|33", 1, "bouteille")]),
    ("coca", [(r"coca", 1, "bouteille")]),
    ("fanta", [(r"fanta", 1, "bouteille")]),
    ("sprite", [(r"sprite", 1, "bouteille")]),
    ("eau", [(r"^eau", 1, "bouteille")]),
    ("jus", [(r"jus", 1, "bouteille")]),
    ("vin", [(r"vin", 1, "bouteille")]),
    ("whisky", [(r"whisky", 1, "bouteille")]),
    ("smoothie", [(r"banane|mangue|fraise|fruit", 0.15, "kg"), (r"sucre", 0.02, "kg"), (r"lait|yaourt", 0.1, "litre")]),
]

# Always-tiny extras that are systematically used.
_DEFAULT_BASE = [(r"sel", 0.002, "kg")]


def _strip_lower(s: str) -> str:
    s = (s or "").strip().lower()
    # Strip common French accents for matching
    repl = {"é": "e", "è": "e", "ê": "e", "ë": "e", "à": "a", "â": "a",
            "ô": "o", "î": "i", "ï": "i", "ç": "c", "ù": "u", "û": "u"}
    for k, v in repl.items():
        s = s.replace(k, v)
    return s


def _compose_ingredients_for_dish(dish_name: str, stock_products: list) -> list:
    """Return a list of {product_id, product_name, quantity, unit} based on dish name keyword scan."""
    import re
    name_norm = _strip_lower(dish_name)
    matched_rules = []
    for keyword, rules in _DISH_KEYWORD_RULES:
        if keyword in name_norm:
            matched_rules.extend(rules)
    # Always include base extras
    matched_rules.extend(_DEFAULT_BASE)
    # Deduplicate by pattern (keep first qty wins)
    seen = set()
    unique_rules = []
    for pat, qty, unit_hint in matched_rules:
        if pat in seen:
            continue
        seen.add(pat)
        unique_rules.append((pat, qty, unit_hint))

    # For each rule, find the best stock product matching the regex
    ingredients = []
    used_product_ids = set()
    for pat, qty, unit_hint in unique_rules:
        regex = re.compile(pat, re.IGNORECASE)
        candidates = [p for p in stock_products if regex.search(_strip_lower(p.get("name", "")))]
        # Prefer products whose unit matches unit_hint
        candidates.sort(key=lambda p: (
            0 if p.get("unit", "").lower() == unit_hint.lower() else 1,
            -float(p.get("quantity", 0) or 0),  # prefer products in stock
        ))
        for c in candidates:
            if c["id"] not in used_product_ids:
                used_product_ids.add(c["id"])
                ingredients.append({
                    "product_id": c["id"],
                    "product_name": c.get("name", ""),
                    "quantity": qty,
                    "unit": c.get("unit") or unit_hint,
                })
                break
    return ingredients


class AutoComposeBody(BaseModel):
    only_unmatched: bool = True  # only generate for caisse products without a recipe yet
    skip_dishless: bool = True   # skip products that match no keyword (drinks already in stock, services, etc.)
    dry_run: bool = False
    selling_price_default: float = 0
    department_filter: Optional[str] = None  # e.g. 'salle_jardin' to only target dishes


@router.post("/recipes/auto-compose")
async def auto_compose_recipes(body: AutoComposeBody):
    """Auto-generate recipes for caisse products by matching dish names to stock products.

    Strategy:
      - For each caisse product (filterable by department), strip-lowercase the name.
      - Apply keyword rules to detect ingredient categories.
      - Pick the best stock product for each rule (prefers matching unit & in-stock items).
      - 1 portion = the conservative default qty in the rule.
      - Save as a stock_recipe with caisse_product_name = caisse_product.name.
    """
    try:
        caisse_products = await db.caisse_products.find({}, {"_id": 0}).to_list(2000)
        if body.department_filter:
            caisse_products = [p for p in caisse_products if p.get("department") == body.department_filter]
        existing_recipes = await db.stock_recipes.find({}, {"_id": 0, "caisse_product_name": 1}).to_list(2000)
        existing_names = {(r.get("caisse_product_name") or "").strip().lower() for r in existing_recipes}
        stock_products = await db.stock_products.find({"is_active": True}, {"_id": 0}).to_list(5000)

        report = {
            "scanned": len(caisse_products),
            "skipped_existing": 0,
            "skipped_no_match": [],
            "created": [],
            "dry_run": body.dry_run,
        }
        new_recipes = []

        for cp in caisse_products:
            cp_name = (cp.get("name") or "").strip()
            if not cp_name:
                continue
            if body.only_unmatched and cp_name.lower() in existing_names:
                report["skipped_existing"] += 1
                continue
            ingredients = _compose_ingredients_for_dish(cp_name, stock_products)
            if not ingredients and body.skip_dishless:
                report["skipped_no_match"].append(cp_name)
                continue
            recipe = {
                "id": str(uuid.uuid4()),
                "name": cp_name,
                "caisse_product_name": cp_name,
                "selling_price": float(cp.get("price") or body.selling_price_default or 0),
                "ingredients": ingredients,
                "notes": "Recette générée automatiquement (1 portion). À ajuster selon vos pratiques.",
                "auto_generated": True,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "updated_at": datetime.now(timezone.utc).isoformat(),
            }
            new_recipes.append(recipe)
            report["created"].append({
                "name": cp_name,
                "ingredients_count": len(ingredients),
                "ingredients": [{"name": i["product_name"], "qty": i["quantity"], "unit": i["unit"]} for i in ingredients],
            })

        if new_recipes and not body.dry_run:
            await db.stock_recipes.insert_many([{**r} for r in new_recipes])

        report["created_count"] = len(report["created"])
        report["skipped_no_match_count"] = len(report["skipped_no_match"])
        return report
    except Exception as e:
        logger.error(f"Auto-compose recipes error: {e}")
        raise HTTPException(500, str(e))



# ==================== REPORTS / RAPPORTS ====================

@router.get("/reports")
async def get_stock_report(
    type: str = "all",  # all, entree, sortie, perte, casse, ajustement
    date_from: str = None,
    date_to: str = None,
    product_id: str = None,
    search: str = None
):
    """Get filtered stock movements report with aggregated stats"""
    query = {}
    
    if type and type != "all":
        if type == "entree":
            query["movement_type"] = {"$in": ["entree", "retour_fournisseur"]}
        elif type == "sortie":
            query["movement_type"] = "sortie"
        elif type == "perte":
            query["movement_type"] = {"$in": ["perte", "casse"]}
        else:
            query["movement_type"] = type
    
    if date_from or date_to:
        dq = {}
        if date_from:
            dq["$gte"] = date_from
        if date_to:
            dq["$lte"] = date_to + "T23:59:59"
        query["created_at"] = dq
    
    if product_id:
        query["product_id"] = product_id
    
    if search:
        query["product_name"] = {"$regex": search, "$options": "i"}
    
    movements = await db.stock_movements.find(query, {"_id": 0}).sort("created_at", -1).to_list(5000)
    
    # Aggregated stats
    total_qty = sum(m.get("quantity", 0) for m in movements)
    total_value = sum(m.get("total_value", 0) for m in movements)
    
    # By type
    by_type = {}
    for m in movements:
        mt = m.get("movement_type", "autre")
        if mt not in by_type:
            by_type[mt] = {"count": 0, "quantity": 0, "value": 0}
        by_type[mt]["count"] += 1
        by_type[mt]["quantity"] += m.get("quantity", 0)
        by_type[mt]["value"] += m.get("total_value", 0)
    
    # By product (top 20)
    by_product = {}
    for m in movements:
        pname = m.get("product_name", "Inconnu")
        if pname not in by_product:
            by_product[pname] = {"count": 0, "quantity": 0, "value": 0}
        by_product[pname]["count"] += 1
        by_product[pname]["quantity"] += m.get("quantity", 0)
        by_product[pname]["value"] += m.get("total_value", 0)
    top_products = sorted(by_product.items(), key=lambda x: x[1]["value"], reverse=True)[:20]
    
    return {
        "movements": movements[:500],
        "total_movements": len(movements),
        "total_quantity": round(total_qty, 2),
        "total_value": round(total_value, 2),
        "by_type": by_type,
        "top_products": [{"name": k, **v} for k, v in top_products],
        "filters": {"type": type, "date_from": date_from, "date_to": date_to, "product_id": product_id, "search": search}
    }

@router.get("/reports/export/pdf")
async def export_report_pdf(
    type: str = "all",
    date_from: str = None,
    date_to: str = None,
    product_id: str = None,
    search: str = None
):
    """Export stock report as PDF"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    
    # Fetch data
    report = await get_stock_report(type=type, date_from=date_from, date_to=date_to, product_id=product_id, search=search)
    movements = report["movements"]
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=15*mm, bottomMargin=15*mm, leftMargin=15*mm, rightMargin=15*mm)
    styles = getSampleStyleSheet()
    elements = []
    
    # Title
    title_style = ParagraphStyle('ReportTitle', parent=styles['Heading1'], fontSize=16, textColor=colors.HexColor('#1a1a2e'), alignment=1)
    elements.append(Paragraph("Rapport de Stock - Espace Maxo", title_style))
    elements.append(Spacer(1, 5*mm))
    
    # Filters info
    type_labels = {"all": "Tous", "entree": "Entrees", "sortie": "Sorties", "perte": "Pertes/Casses", "ajustement": "Ajustements"}
    filter_text = f"Type: {type_labels.get(type, type)}"
    if date_from:
        filter_text += f" | Du: {date_from}"
    if date_to:
        filter_text += f" | Au: {date_to}"
    filter_text += f" | Total: {report['total_movements']} mouvement(s)"
    
    filter_style = ParagraphStyle('FilterInfo', parent=styles['Normal'], fontSize=9, textColor=colors.grey, alignment=1)
    elements.append(Paragraph(filter_text, filter_style))
    elements.append(Spacer(1, 5*mm))
    
    # Summary table
    summary_data = [["Mouvements", "Quantite totale", "Valeur totale"]]
    summary_data.append([
        str(report["total_movements"]),
        f"{report['total_quantity']:.1f}",
        f"{report['total_value']:,.0f} F".replace(",", " ")
    ])
    summary_table = Table(summary_data, colWidths=[60*mm, 60*mm, 60*mm])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2d3748')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f7fafc')),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 8*mm))
    
    # Movements table
    header = ["Date", "Produit", "Type", "Quantite", "P.U.", "Valeur", "Motif"]
    data = [header]
    type_map = {"entree": "Entree", "sortie": "Sortie", "perte": "Perte", "casse": "Casse", "ajustement": "Ajust.", "retour_fournisseur": "Retour", "inventaire": "Inventaire"}
    
    for m in movements[:200]:
        created = m.get("created_at", "")[:16].replace("T", " ")
        data.append([
            created,
            (m.get("product_name", "")[:25]),
            type_map.get(m.get("movement_type", ""), m.get("movement_type", "")),
            f"{m.get('quantity', 0):.2f} {m.get('unit', '')}",
            f"{m.get('unit_price', 0):,.0f}".replace(",", " "),
            f"{m.get('total_value', 0):,.0f}".replace(",", " "),
            (m.get("reason", "")[:30])
        ])
    
    col_w = [30*mm, 40*mm, 18*mm, 25*mm, 20*mm, 22*mm, 35*mm]
    table = Table(data, colWidths=col_w, repeatRows=1)
    
    type_colors = {"Entree": colors.HexColor('#48bb78'), "Sortie": colors.HexColor('#f56565'), "Perte": colors.HexColor('#ed8936'), "Casse": colors.HexColor('#e53e3e')}
    style_commands = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a202c')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, 0), 7),
        ('FONTSIZE', (0, 1), (-1, -1), 6.5),
        ('ALIGN', (3, 0), (5, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#e2e8f0')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f7fafc')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
    ]
    
    for i, row in enumerate(data[1:], 1):
        tc = type_colors.get(row[2])
        if tc:
            style_commands.append(('TEXTCOLOR', (2, i), (2, i), tc))
    
    table.setStyle(TableStyle(style_commands))
    elements.append(table)
    
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"rapport_stock_{type}_{datetime.now(timezone.utc).strftime('%Y%m%d')}.pdf"
    return StreamingResponse(buffer, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename={filename}"})

@router.get("/reports/export/excel")
async def export_report_excel(
    type: str = "all",
    date_from: str = None,
    date_to: str = None,
    product_id: str = None,
    search: str = None
):
    """Export stock report as Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    report = await get_stock_report(type=type, date_from=date_from, date_to=date_to, product_id=product_id, search=search)
    movements = report["movements"]
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mouvements"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="1a202c", end_color="1a202c", fill_type="solid")
    green_font = Font(color="48bb78", bold=True)
    red_font = Font(color="f56565", bold=True)
    orange_font = Font(color="ed8936", bold=True)
    border = Border(
        left=Side(style='thin', color='e2e8f0'), right=Side(style='thin', color='e2e8f0'),
        top=Side(style='thin', color='e2e8f0'), bottom=Side(style='thin', color='e2e8f0')
    )
    
    # Summary row
    type_labels = {"all": "Tous", "entree": "Entrees", "sortie": "Sorties", "perte": "Pertes/Casses", "ajustement": "Ajustements"}
    ws.append(["Rapport de Stock - Espace Maxo"])
    ws.merge_cells('A1:H1')
    ws['A1'].font = Font(bold=True, size=14)
    ws.append([f"Type: {type_labels.get(type, type)}", f"Du: {date_from or 'Debut'}", f"Au: {date_to or 'Fin'}", f"Total: {report['total_movements']} mouvement(s)", f"Valeur: {report['total_value']:,.0f} F"])
    ws.append([])
    
    # Headers
    headers = ["Date", "Produit", "Code", "Type", "Quantite", "Unite", "Prix Unitaire", "Valeur", "Avant", "Apres", "Motif", "Utilisateur"]
    ws.append(headers)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    type_map = {"entree": "Entree", "sortie": "Sortie", "perte": "Perte", "casse": "Casse", "ajustement": "Ajustement", "retour_fournisseur": "Retour", "inventaire": "Inventaire"}
    
    for m in movements:
        created = m.get("created_at", "")[:16].replace("T", " ")
        mt = type_map.get(m.get("movement_type", ""), m.get("movement_type", ""))
        row = [
            created,
            m.get("product_name", ""),
            m.get("product_code", ""),
            mt,
            m.get("quantity", 0),
            m.get("unit", ""),
            m.get("unit_price", 0),
            m.get("total_value", 0),
            m.get("previous_quantity", ""),
            m.get("new_quantity", ""),
            m.get("reason", ""),
            m.get("user_name", "")
        ]
        ws.append(row)
        row_num = ws.max_row
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col).border = border
        # Color the type column
        type_cell = ws.cell(row=row_num, column=4)
        if mt == "Entree":
            type_cell.font = green_font
        elif mt == "Sortie":
            type_cell.font = red_font
        elif mt in ["Perte", "Casse"]:
            type_cell.font = orange_font
    
    # Auto-width columns
    for col_idx, col in enumerate(ws.columns, 1):
        max_len = 0
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        for cell in col:
            try:
                if cell.value and not isinstance(cell, openpyxl.cell.cell.MergedCell):
                    if len(str(cell.value)) > max_len:
                        max_len = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 3, 35)
    
    # Top products sheet
    if report.get("top_products"):
        ws2 = wb.create_sheet("Top Produits")
        ws2.append(["Produit", "Nb mouvements", "Quantite", "Valeur"])
        for col in range(1, 5):
            cell = ws2.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
        for tp in report["top_products"]:
            ws2.append([tp["name"], tp["count"], round(tp["quantity"], 2), round(tp["value"], 2)])
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"rapport_stock_{type}_{datetime.now(timezone.utc).strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={filename}"})

# ==================== SEED DATA ====================

@router.post("/seed")
async def seed_full_data(force: bool = False):
    """Seed the complete restaurant stock database with 25 categories and ~500 products"""
    from routers.stock_data import CATEGORIES, SUPPLIERS, P
    
    existing = await db.stock_products.count_documents({})
    if existing > 0 and not force:
        return {"success": True, "message": f"Donnees deja presentes ({existing} produits). Utilisez force=true pour reinitialiser."}
    
    # Clear existing data if forcing
    if force and existing > 0:
        await db.stock_products.delete_many({})
        await db.stock_categories.delete_many({})
        await db.stock_suppliers.delete_many({})
        await db.stock_movements.delete_many({})
        await db.stock_purchases.delete_many({})
    
    now = datetime.now(timezone.utc).isoformat()
    
    # Create categories
    cat_docs = []
    for c in CATEGORIES:
        cat_docs.append({
            "id": str(uuid.uuid4()),
            "name": c["name"],
            "description": c["description"],
            "color": c["color"],
            "icon": "Package",
            "subcategories": c.get("subcategories", []),
            "created_at": now
        })
    await db.stock_categories.insert_many(cat_docs)
    cat_ids = [c["id"] for c in cat_docs]
    
    # Create suppliers
    sup_docs = []
    for s in SUPPLIERS:
        sup_docs.append({**s, "id": str(uuid.uuid4()), "email": "", "notes": "", "created_at": now})
    await db.stock_suppliers.insert_many(sup_docs)
    
    # Create products
    product_docs = []
    counters = {}
    prefixes = ["CF","LG","FR","LM","VI","VO","PM","PL","OE","HG","EC","BP","SG","BN","BA","CB","EM","EH","GE","AT","PD","SL","SF","FA","SR"]
    for cat_idx, name, unit, smin, price, loc in P:
        pf = prefixes[cat_idx]
        if pf not in counters:
            counters[pf] = 0
        counters[pf] += 1
        code = f"{pf}-{counters[pf]:03d}"
        
        smax = max(smin * 4, 20)
        r = random.random()
        if r < 0.05:
            qty = 0
        elif r < 0.18:
            qty = random.randint(0, max(1, smin - 1))
        else:
            qty = random.randint(smin, smax)
        
        valeur = qty * price
        statut = "rupture" if qty <= 0 else ("faible" if qty <= smin else "normal")
        
        product_docs.append({
            "id": str(uuid.uuid4()), "code": code, "name": name,
            "category_id": cat_ids[cat_idx], "subcategory": "",
            "unit": unit, "quantity": qty, "stock_min": smin, "stock_max": smax,
            "purchase_price": price, "valeur_stock": valeur,
            "supplier_id": "", "storage_location": loc,
            "is_active": True, "photo_url": "",
            "date_achat": "", "date_peremption": "", "observation": "",
            "statut": statut, "created_at": now, "updated_at": now
        })
    
    await db.stock_products.insert_many(product_docs)
    
    return {
        "success": True,
        "message": f"{len(product_docs)} produits, {len(cat_docs)} categories, {len(sup_docs)} fournisseurs crees"
    }

